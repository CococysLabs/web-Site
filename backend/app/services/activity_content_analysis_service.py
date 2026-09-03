"""
Servicio de análisis de contenido para Proyectos, Practicas y Tareas.

Esta es la "Fase 2" del panel de Planeación Curricular. La "Fase 1"
(curriculum_feedback_service) evalúa la coherencia INTERNA de los
documentos de planeación (3_Planeacion_Curricular) y escribe una
observación general por sección en la hoja "Fase 1" de la matriz.

Este servicio hace algo distinto y complementario:

  1. Localiza 3_Planeacion_Curricular y lee de 5_Diseño_Curricular
     qué se PLANEÓ para Proyectos/Practicas/Tareas.
  2. Localiza los documentos REALES entregados en las carpetas
     6_Proyectos / 7_Practicas / 8_Tareas del curso.
  3. Compara ambos con IA y evalúa, campo por campo, las hojas
     "Proyectos" / "Practicas" / "Tareas" de
     0_Revision_de_Material / 02_Matriz observaciones estructura.
  4. Escribe Aplica=Si, Autor=IA, Presente y Observaciones en la
     matriz — ÚNICAMENTE para los ítems que sí tienen contenido
     real en la planeación curricular. Si un ítem no existe en la
     planeación, sus filas no se tocan.
"""
from __future__ import annotations

import io
import json
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill

from app.services.activity_structure_validation_service import (
    activity_structure_validation_service,
)
from app.services.curriculum_feedback_service import (
    MAIN_INPUT_KEY,
    curriculum_feedback_service,
)
from app.services.document_content_validation_service import (
    SUPPORTED_MIMES,
    document_content_validation_service,
)
from app.services.drive_service import drive_service


EXCEL_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

GEMINI_MODEL = "gemini-3.6-flash"

MAX_PLANNING_CHARS = 60_000
MAX_REAL_DOCS_CHARS = 60_000

MATRIX_SEPARATOR = "========================================"


# Tipo de actividad -> hoja de la matriz, clave de planeación y alias de carpeta real.
ACTIVITY_TYPES: List[Dict[str, Any]] = [
    {
        "key": "proyectos",
        "label": "Proyectos",
        "matrix_sheet_names": ("Proyectos",),
        "planning_key": "proyectos",
        "folder_aliases": ["proyectos"],
    },
    {
        "key": "practicas",
        "label": "Practicas",
        "matrix_sheet_names": ("Practicas", "Prácticas"),
        "planning_key": "practicas",
        "folder_aliases": ["practicas", "prácticas"],
    },
    {
        "key": "tareas",
        "label": "Tareas",
        "matrix_sheet_names": ("Tareas",),
        "planning_key": "tareas",
        "folder_aliases": ["tareas"],
    },
]


class ActivityContentAnalysisError(RuntimeError):
    pass


class ActivityContentAnalysisService:

    # ──────────────────────────────────────────────────────────────────────
    # Normalización (reutiliza la misma tolerancia que curriculum_feedback)
    # ──────────────────────────────────────────────────────────────────────

    @staticmethod
    def _normalize(value: object) -> str:
        return curriculum_feedback_service.normalize(value)

    # ──────────────────────────────────────────────────────────────────────
    # Matriz: localizar hoja y parsear bloques por ítem
    # ──────────────────────────────────────────────────────────────────────

    def _find_matrix_sheet(self, wb: "openpyxl.Workbook", names: Tuple[str, ...]):
        wanted = {self._normalize(n) for n in names}
        for title in wb.sheetnames:
            if self._normalize(title) in wanted:
                return wb[title]
        return None

    def _writable_cell(self, ws, row: int, column: int):
        """
        openpyxl solo permite escribir en la celda ancla (superior-izquierda)
        de un rango combinado — las demás son MergedCell de solo lectura.
        Si la celda pedida cae dentro de un rango combinado, retorna la
        celda ancla en su lugar.
        """
        cell = ws.cell(row=row, column=column)

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= column <= merged_range.max_col
                ):
                    return ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col,
                    )

        return cell

    def _parse_matrix_blocks(self, ws) -> List[Dict[str, Any]]:
        """
        Estructura de las hojas Proyectos/Practicas/Tareas:

        Col A = etiqueta del ítem (solo en la primera fila del bloque,
                ej. "Proyecto 1")
        Col B = nombre del campo (ej. "Titulo Proyecto", "Valor", ...)
        Col C = Aplica (Si/No)
        Col D = Autor
        Col E = Presente (Si/No)
        Col F = Observaciones
        """
        blocks: List[Dict[str, Any]] = []
        current: Optional[Dict[str, Any]] = None

        for row in ws.iter_rows(min_row=2, max_col=2):
            a_cell = row[0]
            b_cell = row[1] if len(row) > 1 else None

            a_val = a_cell.value
            b_val = b_cell.value if b_cell is not None else None

            item_label = str(a_val).strip() if a_val not in (None, "") else None
            field_name = str(b_val).strip() if b_val not in (None, "") else None

            if item_label:
                current = {"item_label": item_label, "fields": []}
                blocks.append(current)

            if current is not None and field_name:
                current["fields"].append(
                    {"field_name": field_name, "row_idx": a_cell.row}
                )

        return blocks

    def _apply_ai_result(
        self,
        ws,
        blocks: List[Dict[str, Any]],
        ai_result: Dict[str, Any],
    ) -> Tuple[int, int, int]:
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center = Alignment(horizontal="center", vertical="center")

        items_resp = ai_result.get("items") or {}
        norm_items = {
            self._normalize(k): v
            for k, v in items_resp.items()
            if isinstance(v, dict)
        }

        analyzed = 0
        skipped = 0
        written = 0

        for block in blocks:
            data = norm_items.get(self._normalize(block["item_label"]))

            if not data or not data.get("existe_en_planeacion"):
                skipped += 1
                continue

            analyzed += 1

            fields_resp = data.get("fields") or {}
            norm_fields = {
                self._normalize(k): v
                for k, v in fields_resp.items()
                if isinstance(v, dict)
            }

            for field in block["fields"]:
                f_data = norm_fields.get(self._normalize(field["field_name"]))

                if f_data:
                    presente = (
                        "Si"
                        if str(f_data.get("presente", "No")).strip().lower() == "si"
                        else "No"
                    )
                    observacion = str(f_data.get("observacion", "")).strip()
                else:
                    presente = "No"
                    observacion = "La IA no evaluó este campo."

                row_idx = field["row_idx"]
                fill = green_fill if presente == "Si" else red_fill

                aplica_cell = self._writable_cell(ws, row_idx, 3)
                aplica_cell.value = "Si"
                aplica_cell.alignment = center

                autor_cell = self._writable_cell(ws, row_idx, 4)
                autor_cell.value = "IA"
                autor_cell.alignment = center

                presente_cell = self._writable_cell(ws, row_idx, 5)
                presente_cell.value = presente
                presente_cell.fill = fill
                presente_cell.alignment = center

                obs_cell = self._writable_cell(ws, row_idx, 6)
                obs_cell.value = observacion
                obs_cell.fill = fill
                obs_cell.alignment = wrap

                written += 1

        return analyzed, skipped, written

    # ──────────────────────────────────────────────────────────────────────
    # Planeación curricular: extraer el bloque de una hoja específica
    # ──────────────────────────────────────────────────────────────────────

    def _extract_planning_block(self, full_text: str, planning_key: str) -> str:
        """
        curriculum_feedback_service._extract_curriculum_source() produce un
        texto con bloques delimitados por '===...===' y una línea
        'CLAVE: <clave>' por hoja. Aislamos el bloque de la clave pedida.
        """
        if not full_text:
            return ""

        sections = re.split(
            r"\n\n(?=" + re.escape(MATRIX_SEPARATOR) + r")",
            full_text,
        )

        needle = f"CLAVE: {planning_key}\n"

        for section in sections:
            if needle in section:
                return section.strip()

        return ""

    # ──────────────────────────────────────────────────────────────────────
    # Documentos reales entregados (6_Proyectos / 7_Practicas / 8_Tareas)
    # ──────────────────────────────────────────────────────────────────────

    def _collect_real_documents_text(
        self,
        course_folder_id: str,
        aliases: List[str],
    ) -> Tuple[str, int, Optional[str]]:
        folder = activity_structure_validation_service._find_activity_folder(
            course_folder_id, aliases
        )

        if not folder:
            return "", 0, None

        candidates = activity_structure_validation_service._collect_candidate_files(
            folder["id"]
        )

        parts: List[str] = []
        count = 0

        for f in candidates:
            mime = f.get("mimeType", "")

            if mime not in SUPPORTED_MIMES:
                continue

            content = drive_service.download_file(f["id"])

            if not content:
                continue

            try:
                text = document_content_validation_service.extract_text_from_bytes(
                    content, drive_service.get_effective_mime(mime)
                )
            except Exception:
                continue

            if not text or not text.strip():
                continue

            parts.append(f"--- DOCUMENTO: {f.get('name')} ---\n{text}")
            count += 1

        combined = "\n\n".join(parts)
        return combined[:MAX_REAL_DOCS_CHARS], count, folder["id"]

    # ──────────────────────────────────────────────────────────────────────
    # Gemini
    # ──────────────────────────────────────────────────────────────────────

    def _build_prompt(
        self,
        course_label: str,
        label: str,
        blocks: List[Dict[str, Any]],
        planning_text: str,
        real_docs_text: str,
    ) -> str:
        items_lines: List[str] = []

        for block in blocks:
            items_lines.append(f"- {block['item_label']}:")
            for field in block["fields"]:
                items_lines.append(f"    - {field['field_name']}")

        items_description = "\n".join(items_lines)

        return f"""Eres un evaluador académico universitario experto en coherencia curricular.

CURSO: {course_label}

Tu tarea es analizar los documentos REALES de {label} entregados en el curso,
compararlos con lo que la PLANEACIÓN CURRICULAR del curso dice que debían ser,
y evaluar si cada documento cumple con los campos requeridos por la matriz de
observaciones.

============================================================
REGLAS CRÍTICAS
============================================================

1. La plantilla de planeación puede contener espacios vacíos para {label}
   que el curso NO utilizó (solo etiquetas de la plantilla, sin contenido
   real). Un ítem SOLO existe si la planeación tiene información real y
   sustantiva para él (título, descripción, competencias, tecnologías,
   alcance, etc. con valores reales — no solo el nombre del espacio de la
   plantilla).

2. Para cada ítem listado abajo, determina primero "existe_en_planeacion":
   true si el ítem tiene contenido real en la planeación curricular,
   false si el espacio está vacío o no aparece en absoluto.

3. Si "existe_en_planeacion" es false: para cada uno de sus campos responde
   {{"presente": "No", "observacion": ""}} — estos campos no se escribirán
   en la matriz, solo se piden para completar el JSON correctamente.

4. Si "existe_en_planeacion" es true: evalúa COHERENCIA. El documento real
   entregado, ¿corresponde a lo planeado (mismo tema, competencias,
   alcance)? Si el documento real no existe o no corresponde a lo
   planeado, indícalo explícitamente en las observaciones de sus campos.

5. Para cada campo de la matriz determina "presente": "Si" si el campo
   aparece de forma clara en el documento real (explícita o
   implícitamente); "No" si falta o es insuficiente. En "observacion"
   indica brevemente por qué, y si hay incoherencia con la planeación
   menciónalo ahí.

============================================================
ÍTEMS Y CAMPOS A EVALUAR
============================================================
{items_description}

============================================================
PLANEACIÓN CURRICULAR (lo que el curso planeó para {label})
============================================================
{planning_text or "(No se encontró contenido de planeación para esta sección)"}

============================================================
DOCUMENTOS REALES ENTREGADOS
============================================================
{real_docs_text or "(No se encontraron documentos reales en la carpeta correspondiente)"}

============================================================
FORMATO DE RESPUESTA
============================================================
Responde ÚNICAMENTE con JSON válido, sin markdown ni texto adicional:
{{
  "items": {{
    "<etiqueta exacta del ítem>": {{
      "existe_en_planeacion": true|false,
      "fields": {{
        "<nombre exacto del campo>": {{"presente": "Si"|"No", "observacion": "..."}},
        ...
      }}
    }},
    ...
  }}
}}

Usa EXACTAMENTE las etiquetas de ítem y nombres de campo tal como aparecen
en la sección "ÍTEMS Y CAMPOS A EVALUAR"."""

    def _call_gemini(self, prompt: str) -> Optional[Dict[str, Any]]:
        keys = document_content_validation_service._get_available_keys()

        if not keys:
            return None

        for key in keys:
            short_key = key[:8] + "..."
            backoff = 5.0

            for attempt in range(3):
                try:
                    raw = document_content_validation_service._call_gemini_raw(
                        key, GEMINI_MODEL, prompt
                    )
                    parsed = json.loads(raw)

                    if isinstance(parsed, dict) and isinstance(
                        parsed.get("items"), dict
                    ):
                        return parsed

                    print(
                        "  ⚠️  [Fase 2] Respuesta de Gemini sin 'items' válido"
                    )
                    return None

                except json.JSONDecodeError:
                    print(f"  ⚠️  [Fase 2] JSON inválido de Gemini ({short_key})")
                    break

                except Exception as exc:
                    err = str(exc)
                    is_quota = (
                        "429" in err
                        or "quota" in err.lower()
                        or "RESOURCE_EXHAUSTED" in err
                    )

                    if not is_quota:
                        print(f"  ⚠️  [Fase 2] Error Gemini: {exc}")
                        break

                    daily_exhausted = (
                        "limit: 0" in err or "limit_per_day" in err.lower()
                    )

                    if daily_exhausted:
                        document_content_validation_service._exhausted_keys.add(key)
                        print(
                            f"  🔄 [Fase 2] Key ({short_key}) cuota diaria agotada "
                            "→ rotando"
                        )
                        break

                    if attempt >= 2:
                        document_content_validation_service._exhausted_keys.add(key)
                        print(
                            f"  ⚠️  [Fase 2] Key ({short_key}) no respondió tras "
                            "reintentos → rotando"
                        )
                        break

                    wait = min(backoff, 30.0)
                    print(f"  ⏳ [Fase 2] RPM excedido ({short_key}) — esperando {wait:.0f}s...")
                    time.sleep(wait)
                    backoff *= 2

        return None

    # ──────────────────────────────────────────────────────────────────────
    # Preview (sin IA, sin escribir nada)
    # ──────────────────────────────────────────────────────────────────────

    def preview_course_by_folder(
        self,
        course_folder_id: str,
        course_name: str,
        area: str,
    ) -> Dict[str, Any]:
        """
        Igual que la validación de estructura de Proyectos/Practicas/Tareas:
        el curso se identifica por su carpeta real en Drive (obtenida vía
        Semestre → Área → Curso), no por un código del catálogo de cursos.
        """
        base = {
            "success": False,
            "area": area,
            "name": course_name,
            "folder_id": course_folder_id,
            "error": None,
            "ready_for_analysis": False,
            "ready_for_write": False,
            "matrix_found": False,
            "planning_found": False,
            "activities": {},
        }

        try:
            planning_folder = curriculum_feedback_service._find_planning_folder(
                course_folder_id
            )
        except Exception as exc:
            base["error"] = str(exc)
            return base

        try:
            revision_folder = curriculum_feedback_service._find_revision_folder(
                course_folder_id
            )
            matrix_file = (
                curriculum_feedback_service._find_matrix_file(revision_folder["id"])
                if revision_folder
                else None
            )

            files = curriculum_feedback_service.locate_planning_files(
                planning_folder["id"]
            )
            has_planning = bool(files.get(MAIN_INPUT_KEY))

            activities: Dict[str, Any] = {}

            for spec in ACTIVITY_TYPES:
                folder = activity_structure_validation_service._find_activity_folder(
                    course_folder_id, spec["folder_aliases"]
                )

                doc_count = 0
                if folder:
                    candidates = (
                        activity_structure_validation_service
                        ._collect_candidate_files(folder["id"])
                    )
                    doc_count = len(
                        [f for f in candidates if f.get("mimeType") in SUPPORTED_MIMES]
                    )

                activities[spec["key"]] = {
                    "label": spec["label"],
                    "folder_found": bool(folder),
                    "document_count": doc_count,
                }

            ready_for_analysis = has_planning and any(
                a["folder_found"] for a in activities.values()
            )
            ready_for_write = ready_for_analysis and bool(matrix_file)

            return {
                **base,
                "success": True,
                "ready_for_analysis": ready_for_analysis,
                "ready_for_write": ready_for_write,
                "matrix_found": bool(matrix_file),
                "planning_found": has_planning,
                "activities": activities,
            }

        except Exception as exc:
            base["error"] = str(exc)
            return base

    # ──────────────────────────────────────────────────────────────────────
    # Análisis completo de UN curso
    # ──────────────────────────────────────────────────────────────────────

    def analyze_course_by_folder(
        self,
        course_folder_id: str,
        course_name: str,
        area: str,
        write_output: bool = True,
        activity_keys: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        activity_keys: si se indica, limita el análisis a esos tipos
        ("proyectos", "practicas", "tareas") en vez de los tres —
        reduce la duración de la petición y la carga sobre la IA.
        """
        types_to_process = [
            spec
            for spec in ACTIVITY_TYPES
            if activity_keys is None or spec["key"] in activity_keys
        ]

        if not types_to_process:
            raise ActivityContentAnalysisError(
                "Selecciona al menos un tipo de actividad "
                "(Proyectos, Practicas o Tareas)"
            )

        print(
            "\n"
            "=========================================\n"
            "📋 ANÁLISIS DE CONTENIDO — FASE 2\n"
            f"Curso: {course_name} ({area})\n"
            "Tipos: "
            f"{', '.join(spec['label'] for spec in types_to_process)}\n"
            "=========================================",
            flush=True,
        )

        planning_folder = curriculum_feedback_service._find_planning_folder(
            course_folder_id
        )

        revision_folder = curriculum_feedback_service._find_revision_folder(
            course_folder_id
        )
        if not revision_folder:
            raise ActivityContentAnalysisError(
                "No se encontró 0_Revision_de_Material"
            )

        matrix_file = curriculum_feedback_service._find_matrix_file(
            revision_folder["id"]
        )
        if not matrix_file:
            raise ActivityContentAnalysisError(
                "No se encontró 02_Matriz observaciones estructura"
            )

        files = curriculum_feedback_service.locate_planning_files(
            planning_folder["id"]
        )
        if not files.get(MAIN_INPUT_KEY):
            raise ActivityContentAnalysisError("No se encontró 5_Diseño_Curricular")

        documents, warnings = curriculum_feedback_service.load_context(files)
        full_planning_text = documents[MAIN_INPUT_KEY]["text"]

        matrix_bytes = drive_service.download_file(matrix_file["id"])
        if not matrix_bytes:
            raise ActivityContentAnalysisError(
                "No se pudo descargar la matriz desde Drive"
            )

        wb = openpyxl.load_workbook(io.BytesIO(matrix_bytes))

        activity_results: Dict[str, Any] = {}
        total_written = 0

        for spec in types_to_process:
            print(f"\n  🔎 Procesando: {spec['label']}", flush=True)

            ws = self._find_matrix_sheet(wb, spec["matrix_sheet_names"])
            if ws is None:
                activity_results[spec["key"]] = {
                    "success": False,
                    "error": f"No se encontró la hoja '{spec['label']}' en la matriz",
                }
                continue

            blocks = self._parse_matrix_blocks(ws)
            if not blocks:
                activity_results[spec["key"]] = {
                    "success": False,
                    "error": f"La hoja '{spec['label']}' no tiene filas para evaluar",
                }
                continue

            planning_block = self._extract_planning_block(
                full_planning_text, spec["planning_key"]
            )[:MAX_PLANNING_CHARS]

            real_docs_text, doc_count, _ = self._collect_real_documents_text(
                course_folder_id, spec["folder_aliases"]
            )

            if not planning_block and doc_count == 0:
                activity_results[spec["key"]] = {
                    "success": True,
                    "items_analyzed": 0,
                    "items_skipped": len(blocks),
                    "fields_written": 0,
                    "documents_found": 0,
                    "note": (
                        "Sin planeación ni documentos reales para "
                        f"{spec['label']} — no se modificó la matriz"
                    ),
                }
                continue

            prompt = self._build_prompt(
                f"{course_name} ({area})",
                spec["label"],
                blocks,
                planning_block,
                real_docs_text,
            )

            ai_result = self._call_gemini(prompt)

            if ai_result is None:
                activity_results[spec["key"]] = {
                    "success": False,
                    "error": "La IA no está disponible en este momento",
                    "documents_found": doc_count,
                }
                continue

            analyzed, skipped, written = self._apply_ai_result(
                ws, blocks, ai_result
            )
            total_written += written

            print(
                f"  ✅ {spec['label']}: {analyzed} ítem(s) analizados, "
                f"{skipped} sin planeación, {written} campo(s) escritos",
                flush=True,
            )

            activity_results[spec["key"]] = {
                "success": True,
                "items_analyzed": analyzed,
                "items_skipped": skipped,
                "fields_written": written,
                "documents_found": doc_count,
            }

        matrix_updated = False

        if write_output and total_written > 0:
            try:
                buf = io.BytesIO()
                wb.save(buf)
                matrix_updated = drive_service.upload_file(
                    buf.getvalue(), EXCEL_MIME_TYPE, matrix_file["id"]
                )
            except Exception as exc:
                warnings.append(f"No se pudo guardar la matriz en Drive: {exc}")

        return {
            "success": True,
            "course": {
                "area": area,
                "name": course_name,
                "folder_id": course_folder_id,
            },
            "write_output": write_output,
            "matrix_updated": matrix_updated,
            "fields_written": total_written,
            "activities": activity_results,
            "warnings": warnings,
            "provider": {"name": "gemini", "model": GEMINI_MODEL},
            "matrix": {
                "id": matrix_file.get("id"),
                "name": matrix_file.get("name"),
                "webViewLink": matrix_file.get("webViewLink"),
            },
        }


# Instancia global
activity_content_analysis_service = ActivityContentAnalysisService()
