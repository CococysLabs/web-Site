from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.models.course_catalog import CourseCatalog
from app.services.drive_service import drive_service
from time import perf_counter

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_SHEET_ID = (
    "1CRB-15gwBXxS1NmfFFgJaDqWeBSkHqaDK9JhdsRcEbI"
)

DEFAULT_ROOT_ID = (
    "1kKtxjCV9cXxkS_BeQv95Ud5M_Q0S77aA"
)

SEMESTERS = {
    "1S": (
        "Primer Semestre",
        "Primer_Semestre",
    ),
    "2S": (
        "Segundo Semestre",
        "Segundo_Semestre",
    ),
}

AREA_FOLDER_NAMES = {
    "Software": "2. Software",
    "Sistemas": "3. Sistemas",
    "Computacion": "4. Computacion",
}

COLUMN_ALIASES = {
    "area": {
        "area",
    },
    "code": {
        "codigo",
        "codigo curso",
        "codigo del curso",
    },
    "course": {
        "curso",
        "nombre curso",
        "nombre del curso",
    },
    "section": {
        "seccion",
    },
    "name": {
        "nombre",
        "nombre completo",
    },
    "frequent_email": {
        "correofrecuente",
        "correo frecuente",
        "correo de uso frecuente",
    },
    "institutional_email": {
        "correoinstitucional",
        "correo institucional",
    },
    "role": {
        "funcion",
        "cargo",
    },
    "phone": {
        "telefono",
        "numero de telefono",
    },
    "semester_start": {
        "semestre inicial",
        "semestreinicio",
    },
    "semester_end": {
        "semestre final",
        "semestrefinal",
    },
}

# Límites para evitar recorrer rangos enormes generados por
# formatos residuales dentro de Google Sheets.
MAX_HEADER_SCAN_ROWS = 15
MAX_HEADER_SCAN_COLUMNS = 50
MAX_SOURCE_DATA_ROWS = 5000
STOP_AFTER_EMPTY_ROWS = 100

class CourseContactsError(RuntimeError):
    """
    Error controlado del módulo de contactos.
    """


@dataclass(frozen=True)
class Contact:
    area: str
    code: str
    course: str
    section: str
    name: str
    frequent_email: str
    institutional_email: str
    role: str
    phone: str
    semester_start: str
    semester_end: str


@dataclass(frozen=True)
class ContactsSource:
    records: List[Contact]
    spreadsheet_name: str
    docentes_sheet: str
    auxiliares_sheet: str
    warnings: List[str]


class CourseContactsService:
    """
    Lee docentes y auxiliares desde el Google Sheets maestro.

    Después genera un archivo XLSX por curso y lo guarda dentro de:

    año_semestre/curso/2_Contactos
    """

    @staticmethod
    def normalize(value: object) -> str:
        """
        Normalizar textos para compararlos sin acentos ni diferencias
        entre espacios, guiones o guiones bajos.
        """
        text = str(value or "").strip().lower()
        text = text.replace("_", " ")

        text = unicodedata.normalize("NFKD", text)
        text = "".join(
            character
            for character in text
            if not unicodedata.combining(character)
        )

        text = re.sub(r"[^a-z0-9]+", " ", text)
        return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def clean(value: object) -> str:
        """
        Convertir valores de Excel en texto.

        Evita que un código como 774 sea convertido en 774.0.
        """
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    def canonical_area(self, value: object) -> str:
        """
        Convierte:
        - Computación -> Computacion
        - 4. Computacion -> Computacion
        """
        normalized = self.normalize(value)

        normalized = re.sub(
            r"^\d+\s*",
            "",
            normalized,
        ).replace(" ", "")

        areas = {
            "computacion": "Computacion",
            "sistemas": "Sistemas",
            "software": "Software",
        }

        return areas.get(
            normalized,
            self.clean(value),
        )
        
    def area_folder_name(self, area: str) -> str:
        """
        Convertir el área del catálogo en el nombre real de su carpeta de Drive.

        Ejemplos:
        - Software -> 2. Software
        - Sistemas -> 3. Sistemas
        - Computacion -> 4. Computacion
        """
        canonical = self.canonical_area(area)

        folder_name = AREA_FOLDER_NAMES.get(canonical)

        if not folder_name:
            raise CourseContactsError(
                f"No existe una carpeta configurada para el área '{area}'"
            )

        return folder_name

    def normalize_semester(self, value: str) -> str:
        normalized = self.normalize(value).replace(" ", "")

        aliases = {
            "1s": "1S",
            "primersemestre": "1S",
            "primero": "1S",
            "2s": "2S",
            "segundosemestre": "2S",
            "segundo": "2S",
        }

        code = aliases.get(normalized)

        if not code:
            raise CourseContactsError(
                "El semestre debe ser 1S o 2S"
            )

        return code

    def period_data(
        self,
        semester: str,
        year: int,
    ) -> Dict[str, str]:
        """
        Construye todos los nombres dependientes del periodo.

        Para 2S y 2026:

        carpeta:
            2026_Segundo_Semestre

        hojas:
            Docentes_2S_2026_
            Auxiliares_2S_2026_
        """
        code = self.normalize_semester(semester)
        label, folder_slug = SEMESTERS[code]

        return {
            "code": code,
            "label": label,
            "folder": f"{year}_{folder_slug}",
            "docentes": f"Docentes_{code}_{year}_",
            "auxiliares": f"Auxiliares_{code}_{year}_",
            "title": f"Lista de Contactos {label} {year}",
        }

    def contacts_filename(
        self,
        course: CourseCatalog,
        semester: str,
        year: int,
    ) -> str:
        """
        Construir el nombre exacto del archivo manual de contactos.

        Ejemplos:
        - 774_Contactos_1S_2026.xlsx
        - 774_Contactos_2S_2026.xlsx
        """
        semester_code = self.normalize_semester(
            semester
        )

        course_code = self.clean(
            course.code
        )

        return (
            f"{course_code}_Contactos_"
            f"{semester_code}_{year}.xlsx"
        )

    @staticmethod
    def spreadsheet_id() -> str:
        return (
            getattr(
                settings,
                "GOOGLE_CONTACTS_SPREADSHEET_ID",
                None,
            )
            or DEFAULT_SHEET_ID
        )

    @staticmethod
    def root_folder_id() -> str:
        return (
            getattr(
                settings,
                "GOOGLE_DRIVE_STRUCTURE_FOLDER_ID",
                None,
            )
            or DEFAULT_ROOT_ID
        )

    @staticmethod
    def image_path() -> Path:
        """
        Ruta predeterminada:

        backend/app/static/images/contactos-buho.png
        """
        configured = getattr(
            settings,
            "GOOGLE_CONTACTS_IMAGE_PATH",
            None,
        )

        if configured:
            path = Path(configured)

            if path.is_absolute():
                return path

            return Path.cwd() / path

        return (
            Path(__file__).resolve().parents[1]
            / "static"
            / "images"
            / "contactos-buho.png"
        )

    def get_sheet(
        self,
        workbook,
        expected_name: str,
    ):
        """
        Buscar una hoja por nombre exacto y, como respaldo,
        por nombre normalizado.
        """
        if expected_name in workbook.sheetnames:
            return workbook[expected_name]

        normalized_expected = self.normalize(expected_name)

        for sheet_name in workbook.sheetnames:
            if self.normalize(sheet_name) == normalized_expected:
                return workbook[sheet_name]

        raise CourseContactsError(
            f"No se encontró la hoja '{expected_name}'. "
            f"Hojas disponibles: {workbook.sheetnames}"
        )

    def header_map(
    self,
    worksheet,
) -> Tuple[int, Dict[str, int]]:
        """
        Detectar la fila de encabezados sin recorrer todas las columnas
        declaradas por el XLSX.

        Google Sheets puede exportar hojas con dimensiones mayores que
        el contenido real debido a formatos residuales.
        """
        started_at = perf_counter()

        alias_lookup = {
            self.normalize(alias): canonical
            for canonical, aliases in COLUMN_ALIASES.items()
            for alias in aliases
        }

        declared_columns = (
            worksheet.max_column
            or MAX_HEADER_SCAN_COLUMNS
        )

        maximum_columns = min(
            declared_columns,
            MAX_HEADER_SCAN_COLUMNS,
        )

        print(
            "👥 [CONTACTOS] Buscando encabezados | "
            f"hoja={worksheet.title} | "
            f"max_column_declarado={worksheet.max_column} | "
            f"columnas_revisadas={maximum_columns}",
            flush=True,
        )

        rows = worksheet.iter_rows(
            min_row=1,
            max_row=MAX_HEADER_SCAN_ROWS,
            min_col=1,
            max_col=maximum_columns,
            values_only=True,
        )

        for row_number, row in enumerate(
            rows,
            start=1,
        ):
            found_columns: Dict[str, int] = {}

            for column_number, raw_value in enumerate(
                row,
                start=1,
            ):
                normalized_value = self.normalize(
                    raw_value
                )

                canonical = alias_lookup.get(
                    normalized_value
                )

                if (
                    canonical
                    and canonical not in found_columns
                ):
                    found_columns[canonical] = (
                        column_number
                    )

            required_columns = {
                "area",
                "code",
                "course",
                "name",
            }

            if required_columns.issubset(
                found_columns
            ):
                print(
                    "✅ [CONTACTOS] Encabezados encontrados | "
                    f"hoja={worksheet.title} | "
                    f"fila={row_number} | "
                    f"columnas={found_columns} | "
                    f"segundos={perf_counter() - started_at:.2f}",
                    flush=True,
                )

                return row_number, found_columns

        raise CourseContactsError(
            f"No se reconocieron los encabezados "
            f"de la hoja '{worksheet.title}' "
            f"dentro de las primeras "
            f"{MAX_HEADER_SCAN_ROWS} filas"
        )
    
    def row_value(
        self,
        row: Sequence[object],
        columns: Dict[str, int],
        field: str,
    ) -> str:
        column_number = columns.get(field)

        if (
            not column_number
            or column_number > len(row)
        ):
            return ""

        return self.clean(
            row[column_number - 1]
        )

    def parse_sheet(
        self,
        worksheet,
        default_role: str,
        require_semester_range: bool = False,
    ) -> Tuple[List[Contact], List[str]]:
        """
        Leer contactos de una hoja sin recorrer rangos vacíos enormes.

        Docentes:
        - No requieren Semestre Inicial ni Semestre Final.
        - Ambos valores quedan vacíos.

        Auxiliares:
        - Sí deben tener Semestre Inicial y Semestre Final.

        La lectura se detiene al encontrar 100 filas principales
        consecutivas completamente vacías.
        """
        started_at = perf_counter()

        header_row, columns = self.header_map(
            worksheet
        )

        contacts: List[Contact] = []
        warnings: List[str] = []

        missing_semester_columns = []

        if "semester_start" not in columns:
            missing_semester_columns.append(
                "Semestre Inicial"
            )

        if "semester_end" not in columns:
            missing_semester_columns.append(
                "Semestre Final"
            )

        if (
            require_semester_range
            and missing_semester_columns
        ):
            warnings.append(
                f"La hoja '{worksheet.title}' no contiene "
                f"las columnas requeridas: "
                f"{', '.join(missing_semester_columns)}."
            )

        first_data_row = header_row + 1

        last_data_row = (
            first_data_row
            + MAX_SOURCE_DATA_ROWS
            - 1
        )

        maximum_column = min(
            max(columns.values()),
            MAX_HEADER_SCAN_COLUMNS,
        )

        print(
            "👥 [CONTACTOS] Iniciando lectura de hoja | "
            f"hoja={worksheet.title} | "
            f"fila_inicial={first_data_row} | "
            f"fila_limite={last_data_row} | "
            f"columnas={maximum_column}",
            flush=True,
        )

        rows = worksheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=1,
            max_col=maximum_column,
            values_only=True,
        )

        empty_rows = 0
        processed_rows = 0
        incomplete_rows = 0

        for row_number, row in enumerate(
            rows,
            start=first_data_row,
        ):
            processed_rows += 1

            area = self.row_value(
                row,
                columns,
                "area",
            )

            code = self.row_value(
                row,
                columns,
                "code",
            )

            course_name = self.row_value(
                row,
                columns,
                "course",
            )

            name = self.row_value(
                row,
                columns,
                "name",
            )

            principal_values = [
                area,
                code,
                course_name,
                name,
            ]

            if not any(principal_values):
                empty_rows += 1

                if empty_rows >= STOP_AFTER_EMPTY_ROWS:
                    print(
                        "ℹ️ [CONTACTOS] Lectura detenida por "
                        "filas vacías consecutivas | "
                        f"hoja={worksheet.title} | "
                        f"fila={row_number} | "
                        f"filas_vacias={empty_rows}",
                        flush=True,
                    )
                    break

                continue

            # Reiniciar porque encontramos una fila con contenido.
            empty_rows = 0

            if not code or not name:
                incomplete_rows += 1

                # Evitar generar miles de advertencias si la hoja
                # contiene datos residuales.
                if incomplete_rows <= 50:
                    warnings.append(
                        f"Fila {row_number} de "
                        f"'{worksheet.title}' omitida: "
                        "falta código o nombre."
                    )

                continue

            role = (
                self.row_value(
                    row,
                    columns,
                    "role",
                )
                or default_role
            )

            contacts.append(
                Contact(
                    area=area,
                    code=code,
                    course=course_name,
                    section=self.row_value(
                        row,
                        columns,
                        "section",
                    ),
                    name=name,
                    frequent_email=self.row_value(
                        row,
                        columns,
                        "frequent_email",
                    ),
                    institutional_email=self.row_value(
                        row,
                        columns,
                        "institutional_email",
                    ),
                    role=role,
                    phone=self.row_value(
                        row,
                        columns,
                        "phone",
                    ),
                    semester_start=self.row_value(
                        row,
                        columns,
                        "semester_start",
                    ),
                    semester_end=self.row_value(
                        row,
                        columns,
                        "semester_end",
                    ),
                )
            )

        if incomplete_rows > 50:
            warnings.append(
                f"La hoja '{worksheet.title}' contiene "
                f"{incomplete_rows} filas incompletas. "
                "Solo se muestran las primeras 50 advertencias."
            )

        print(
            "✅ [CONTACTOS] Hoja procesada | "
            f"hoja={worksheet.title} | "
            f"filas_revisadas={processed_rows} | "
            f"contactos={len(contacts)} | "
            f"filas_incompletas={incomplete_rows} | "
            f"segundos={perf_counter() - started_at:.2f}",
            flush=True,
        )

        return contacts, warnings

    def load_source(
        self,
        semester: str,
        year: int,
    ) -> ContactsSource:
        """
        Descargar el Google Sheets como XLSX y leer sus dos hojas.

        No usa Google Sheets API. Usa Drive API porque drive_service
        exporta archivos application/vnd.google-apps.spreadsheet
        a XLSX.

        El libro se abre en modo read_only para reducir el consumo
        de memoria dentro de Render.
        """
        started_at = perf_counter()
        workbook = None
        content = None

        print(
            "👥 [CONTACTOS] Iniciando carga de fuente | "
            f"semestre={semester} | "
            f"año={year}",
            flush=True,
        )

        if not drive_service.service:
            raise CourseContactsError(
                "El servicio de Google Drive no está inicializado"
            )

        period = self.period_data(
            semester,
            year,
        )

        spreadsheet_id = self.spreadsheet_id()

        metadata_started_at = perf_counter()

        metadata = drive_service.get_file_metadata(
            spreadsheet_id
        )

        print(
            "👥 [CONTACTOS] Metadatos de fuente terminados | "
            f"segundos={perf_counter() - metadata_started_at:.2f} | "
            f"encontrado={bool(metadata)}",
            flush=True,
        )

        if not metadata:
            raise CourseContactsError(
                "No se pudo acceder al Google Sheets de contactos. "
                "Compártelo con la cuenta de servicio."
            )

        download_started_at = perf_counter()

        content = drive_service.download_file(
            spreadsheet_id
        )

        print(
            "👥 [CONTACTOS] Descarga de fuente terminada | "
            f"segundos={perf_counter() - download_started_at:.2f} | "
            f"bytes={len(content) if content else 0}",
            flush=True,
        )

        if not content:
            raise CourseContactsError(
                "No se pudo exportar el Google Sheets "
                "de contactos como XLSX"
            )

        try:
            open_started_at = perf_counter()

            print(
                "👥 [CONTACTOS] Abriendo XLSX con openpyxl | "
                f"bytes={len(content)} | "
                "modo=read_only",
                flush=True,
            )

            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )

            print(
                "✅ [CONTACTOS] XLSX abierto | "
                f"segundos={perf_counter() - open_started_at:.2f} | "
                f"hojas={workbook.sheetnames}",
                flush=True,
            )

            docentes_worksheet = self.get_sheet(
                workbook,
                period["docentes"],
            )

            auxiliares_worksheet = self.get_sheet(
                workbook,
                period["auxiliares"],
            )

            docentes_started_at = perf_counter()

            docentes, docentes_warnings = self.parse_sheet(
                worksheet=docentes_worksheet,
                default_role="Docente",
                require_semester_range=False,
            )

            print(
                "✅ [CONTACTOS] Docentes terminados | "
                f"cantidad={len(docentes)} | "
                f"segundos={perf_counter() - docentes_started_at:.2f}",
                flush=True,
            )

            auxiliares_started_at = perf_counter()

            auxiliares, auxiliares_warnings = self.parse_sheet(
                worksheet=auxiliares_worksheet,
                default_role="Auxiliar",
                require_semester_range=True,
            )

            print(
                "✅ [CONTACTOS] Auxiliares terminados | "
                f"cantidad={len(auxiliares)} | "
                f"segundos={perf_counter() - auxiliares_started_at:.2f}",
                flush=True,
            )

            # Eliminar filas duplicadas exactas.
            # Secciones diferentes no se consideran duplicadas.
            unique_contacts: Dict[
                Tuple[str, ...],
                Contact,
            ] = {}

            for contact in docentes + auxiliares:
                key = (
                    self.canonical_area(
                        contact.area
                    ),
                    contact.code,
                    self.normalize(
                        contact.name
                    ),
                    self.normalize(
                        contact.role
                    ),
                    self.normalize(
                        contact.section
                    ),
                )

                unique_contacts[key] = contact

            result = ContactsSource(
                records=list(
                    unique_contacts.values()
                ),
                spreadsheet_name=metadata.get(
                    "name",
                    "Contactos",
                ),
                docentes_sheet=docentes_worksheet.title,
                auxiliares_sheet=auxiliares_worksheet.title,
                warnings=(
                    docentes_warnings
                    + auxiliares_warnings
                ),
            )

            print(
                "✅ [CONTACTOS] Fuente procesada | "
                f"docentes={len(docentes)} | "
                f"auxiliares={len(auxiliares)} | "
                f"contactos_unicos={len(result.records)} | "
                f"total_segundos={perf_counter() - started_at:.2f}",
                flush=True,
            )

            return result

        except CourseContactsError:
            raise

        except Exception as exception:
            print(
                "❌ [CONTACTOS] Error procesando XLSX | "
                f"tipo={type(exception).__name__} | "
                f"mensaje={exception} | "
                f"total_segundos={perf_counter() - started_at:.2f}",
                flush=True,
            )

            raise CourseContactsError(
                f"No se pudo abrir o procesar el archivo maestro "
                f"de contactos: {exception}"
            ) from exception

        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception as close_exception:
                    print(
                        "⚠️ [CONTACTOS] No se pudo cerrar el workbook | "
                        f"mensaje={close_exception}",
                        flush=True,
                    )

            # Liberar la referencia al contenido XLSX.
            content = None

    def contacts_for_course(
        self,
        records: Iterable[Contact],
        course: CourseCatalog,
    ) -> List[Contact]:
        """
        Filtrar por área y código de curso.
        """
        selected_area = self.canonical_area(
            course.area
        )

        selected_code = self.clean(
            course.code
        )

        contacts = [
            contact
            for contact in records
            if (
                self.canonical_area(contact.area)
                == selected_area
                and self.clean(contact.code)
                == selected_code
            )
        ]

        def sort_key(contact: Contact):
            role_order = (
                0
                if self.normalize(contact.role) == "docente"
                else 1
            )

            return (
                role_order,
                self.normalize(contact.section),
                self.normalize(contact.name),
            )

        return sorted(
            contacts,
            key=sort_key,
        )

    def find_period_folder(
        self,
        semester: str,
        year: int,
    ) -> Dict:
        """
        Obtener la carpeta del período.

        En la configuración actual:
        GOOGLE_DRIVE_STRUCTURE_FOLDER_ID apunta directamente a:

            2026_Segundo_Semestre

        Por lo tanto, primero se verifica si la carpeta raíz configurada
        ya es la carpeta del período.

        También se mantiene compatibilidad por si en el futuro el ID apunta
        a una carpeta superior que contiene varios períodos.
        """
        period = self.period_data(
            semester,
            year,
        )

        root_folder_id = self.root_folder_id()

        root_metadata = drive_service.get_file_metadata(
            root_folder_id
        )

        if not root_metadata:
            raise CourseContactsError(
                "No se pudo acceder a "
                "GOOGLE_DRIVE_STRUCTURE_FOLDER_ID. "
                "Verifica el ID y los permisos de la service account."
            )

        root_name = self.clean(
            root_metadata.get("name")
        )

        expected_period_name = period["folder"]

        # Caso actual:
        # GOOGLE_DRIVE_STRUCTURE_FOLDER_ID ya es 2026_Segundo_Semestre.
        if (
            self.normalize(root_name)
            == self.normalize(expected_period_name)
        ):
            return {
                **root_metadata,
                "id": (
                    root_metadata.get("id")
                    or root_folder_id
                ),
                "name": (
                    root_metadata.get("name")
                    or expected_period_name
                ),
            }

        # Compatibilidad futura:
        # el ID podría apuntar a una carpeta padre que contiene los períodos.
        folder = drive_service.find_folder(
            expected_period_name,
            root_folder_id,
        )

        if folder:
            return folder

        raise CourseContactsError(
            f"La carpeta configurada como raíz se llama "
            f"'{root_name}', pero se esperaba "
            f"'{expected_period_name}'. Tampoco se encontró "
            "una carpeta de período dentro de ella."
        )
    
    def find_area_folder(
        self,
        period_folder_id: str,
        area: str,
    ) -> Dict:
        """
        Buscar la carpeta del área dentro de la carpeta del período.

        Ejemplo:
        2026_Segundo_Semestre/3. Sistemas
        """
        expected_name = self.area_folder_name(area)

        area_folder = drive_service.find_folder(
            expected_name,
            period_folder_id,
        )

        if not area_folder:
            raise CourseContactsError(
                f"No se encontró la carpeta de área '{expected_name}' "
                "dentro de la carpeta del período"
            )

        return area_folder

    def find_course_folder(
        self,
        area_folder_id: str,
        course: CourseCatalog,
        semester: str,
        year: int,
    ) -> Optional[Dict]:
        """
        Buscar la carpeta del curso dentro de la carpeta de su área.

        Ruta:
        período/área/curso
        """
        code = re.escape(
            self.clean(course.code)
        )

        pattern = re.compile(
            rf"^{code}(?:[_\s.\-]|$)",
            re.IGNORECASE,
        )

        candidates = [
            folder
            for folder in drive_service.list_folders(
                area_folder_id
            )
            if pattern.search(
                self.clean(
                    folder.get("name")
                )
            )
        ]

        if len(candidates) == 1:
            return candidates[0]

        period_code = self.normalize_semester(
            semester
        ).lower()

        preferred = [
            folder
            for folder in candidates
            if (
                period_code
                in self.normalize(
                    folder.get("name")
                ).replace(" ", "")
                and str(year)
                in self.clean(
                    folder.get("name")
                )
            )
        ]

        if len(preferred) == 1:
            return preferred[0]

        return None

    def find_contacts_folder(
        self,
        course_folder_id: str,
    ) -> Optional[Dict]:
        folders = drive_service.list_folders(
            course_folder_id
        )

        # Primero buscar el nombre exacto.
        for folder in folders:
            if folder.get("name") == "2_Contactos":
                return folder

        # Respaldo para diferencias menores.
        for folder in folders:
            normalized = self.normalize(
                folder.get("name")
            )

            if normalized in {
                "2 contactos",
                "contactos",
            }:
                return folder

        return None

    def build_excel(
        self,
        course: CourseCatalog,
        contacts: Sequence,
        semester: str,
        year: int,
    ) -> bytes:
        """
        Generar el archivo Contactos.xlsx con:

        - Encabezado compacto.
        - Celdas y cuadrícula visibles.
        - Sin filas ni columnas congeladas.
        - Tabla desde la columna A.
        - Bordes en todas las celdas.
        - Filas alternadas para mejorar la lectura.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Contactos"

        # ---------------------------------------------------------
        # Configuración de la hoja
        # ---------------------------------------------------------

        # Mostrar las líneas normales de las celdas.
        worksheet.sheet_view.showGridLines = True

        # No congelar ninguna fila ni columna.
        worksheet.freeze_panes = None
        worksheet.sheet_view.pane = None

        worksheet.sheet_view.zoomScale = 90

        semester_code = str(semester or "").strip().upper()

        semester_label = {
            "1S": "Primer Semestre",
            "2S": "Segundo Semestre",
        }.get(
            semester_code,
            str(semester or "").strip(),
        )

        title = (
            f"Lista de Contactos "
            f"{semester_label} {year}"
        )

        # ---------------------------------------------------------
        # Estilos
        # ---------------------------------------------------------

        orange = "F25C05"
        light_orange = "FCE4D6"
        metadata_fill = "FFF2CC"
        header_fill = "BFBFBF"
        alternate_fill = "F2F2F2"
        white_fill = "FFFFFF"
        border_color = "7F7F7F"

        thin_side = Side(
            style="thin",
            color=border_color,
        )

        medium_side = Side(
            style="medium",
            color="595959",
        )

        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        header_border = Border(
            left=thin_side,
            right=thin_side,
            top=medium_side,
            bottom=medium_side,
        )

        centered = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        left_aligned = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        # ---------------------------------------------------------
        # Tamaño de columnas
        # ---------------------------------------------------------

        column_widths = {
            "A": 35,  # Nombre
            "B": 14,  # Función
            "C": 12,  # Sección
            "D": 32,  # Correo frecuente
            "E": 40,  # Correo institucional
            "F": 16,  # Teléfono
            "G": 24,  # Semestre inicial
            "H": 24,  # Semestre final
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        # ---------------------------------------------------------
        # Encabezado superior compacto
        # ---------------------------------------------------------

        # Logo en A1.
        worksheet["A1"] = ""
        worksheet["A1"].fill = PatternFill(
            "solid",
            fgColor=light_orange,
        )
        worksheet["A1"].border = thin_border

        # Título solamente en una fila.
        worksheet.merge_cells("B1:H1")

        title_cell = worksheet["B1"]
        title_cell.value = title
        title_cell.font = Font(
            name="Arial",
            size=16,
            bold=True,
            color=orange,
        )
        title_cell.fill = PatternFill(
            "solid",
            fgColor=light_orange,
        )
        title_cell.alignment = centered
        title_cell.border = thin_border

        # Aplicar relleno y borde a todo el rango del título.
        for column_number in range(2, 9):
            cell = worksheet.cell(
                row=1,
                column=column_number,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=light_orange,
            )
            cell.border = thin_border

        worksheet.row_dimensions[1].height = 30

        # ---------------------------------------------------------
        # Datos del curso
        # ---------------------------------------------------------

        worksheet["B2"] = "Nombre del Curso"
        worksheet["B2"].font = Font(
            name="Arial",
            size=10,
            bold=True,
        )
        worksheet["B2"].fill = PatternFill(
            "solid",
            fgColor=metadata_fill,
        )
        worksheet["B2"].alignment = left_aligned
        worksheet["B2"].border = thin_border

        worksheet.merge_cells("C2:E2")
        worksheet["C2"] = str(
            course.name or ""
        )
        worksheet["C2"].font = Font(
            name="Arial",
            size=10,
        )
        worksheet["C2"].fill = PatternFill(
            "solid",
            fgColor=white_fill,
        )
        worksheet["C2"].alignment = left_aligned
        worksheet["C2"].border = thin_border

        worksheet["F2"] = "Código del Curso"
        worksheet["F2"].font = Font(
            name="Arial",
            size=10,
            bold=True,
        )
        worksheet["F2"].fill = PatternFill(
            "solid",
            fgColor=metadata_fill,
        )
        worksheet["F2"].alignment = left_aligned
        worksheet["F2"].border = thin_border

        worksheet.merge_cells("G2:H2")
        worksheet["G2"] = str(
            course.code or ""
        )
        worksheet["G2"].font = Font(
            name="Arial",
            size=10,
        )
        worksheet["G2"].fill = PatternFill(
            "solid",
            fgColor=white_fill,
        )
        worksheet["G2"].alignment = centered
        worksheet["G2"].border = thin_border

        # Aplicar bordes al rango completo B2:H2.
        for column_number in range(2, 9):
            worksheet.cell(
                row=2,
                column=column_number,
            ).border = thin_border

        worksheet.row_dimensions[2].height = 24

        # ---------------------------------------------------------
        # Resumen de contactos
        # ---------------------------------------------------------

        docentes_count = 0
        auxiliares_count = 0

        for contact in contacts:
            role = str(
                getattr(
                    contact,
                    "role",
                    "",
                )
                or ""
            ).strip().lower()

            if role == "docente":
                docentes_count += 1
            elif role == "auxiliar":
                auxiliares_count += 1

        summary_values = [
            "Total de Contactos",
            len(contacts),
            "Docentes",
            docentes_count,
            "Auxiliares",
            auxiliares_count,
            "Período",
            f"{semester_code} {year}",
        ]

        for column_number, value in enumerate(
            summary_values,
            start=1,
        ):
            cell = worksheet.cell(
                row=3,
                column=column_number,
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = centered

            if column_number in {
                1,
                3,
                5,
                7,
            }:
                cell.font = Font(
                    name="Arial",
                    size=9,
                    bold=True,
                )
                cell.fill = PatternFill(
                    "solid",
                    fgColor=metadata_fill,
                )
            else:
                cell.font = Font(
                    name="Arial",
                    size=9,
                )
                cell.fill = PatternFill(
                    "solid",
                    fgColor=white_fill,
                )

        worksheet.row_dimensions[3].height = 22

        # Fila pequeña para separar encabezado y tabla.
        worksheet.row_dimensions[4].height = 7

        # ---------------------------------------------------------
        # Imagen
        # ---------------------------------------------------------

        configured_image_path = getattr(
            settings,
            "GOOGLE_CONTACTS_IMAGE_PATH",
            None,
        )

        if configured_image_path:
            image_path = Path(
                configured_image_path
            )

            if not image_path.is_absolute():
                image_path = (
                    Path.cwd()
                    / image_path
                )
        else:
            image_path = (
                Path(__file__).resolve().parents[1]
                / "static"
                / "images"
                / "contactos-buho.png"
            )

        if image_path.exists():
            try:
                logo = OpenpyxlImage(
                    str(image_path)
                )
                logo.width = 68
                logo.height = 68
                logo.anchor = "A1"

                worksheet.add_image(
                    logo
                )
            except Exception as exception:
                print(
                    "⚠️ No se pudo agregar la imagen "
                    f"al Excel: {exception}"
                )
        else:
            print(
                "⚠️ No se encontró la imagen de contactos: "
                f"{image_path}"
            )

        # ---------------------------------------------------------
        # Encabezados de la tabla
        # ---------------------------------------------------------

        header_row = 5

        headers = [
            "Nombre",
            "Función",
            "Sección",
            "Correo de Uso Frecuente",
            "Correo Institucional",
            "Teléfono",
            "Semestre Inicial",
            "Semestre Final",
        ]

        for column_number, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_number,
            )

            cell.value = header
            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color="000000",
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=header_fill,
            )
            cell.alignment = centered
            cell.border = header_border

        worksheet.row_dimensions[
            header_row
        ].height = 32

        # ---------------------------------------------------------
        # Filas de contactos
        # ---------------------------------------------------------

        first_data_row = header_row + 1

        for row_offset, contact in enumerate(
            contacts,
        ):
            row_number = (
                first_data_row
                + row_offset
            )

            values = [
                getattr(
                    contact,
                    "name",
                    "",
                ),
                getattr(
                    contact,
                    "role",
                    "",
                ),
                getattr(
                    contact,
                    "section",
                    "",
                ),
                getattr(
                    contact,
                    "frequent_email",
                    "",
                ),
                getattr(
                    contact,
                    "institutional_email",
                    "",
                ),
                getattr(
                    contact,
                    "phone",
                    "",
                ),
                getattr(
                    contact,
                    "semester_start",
                    "",
                ),
                getattr(
                    contact,
                    "semester_end",
                    "",
                ),
            ]

            row_fill = (
                alternate_fill
                if row_offset % 2
                else white_fill
            )

            for column_number, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                cell.value = (
                    "" if value is None
                    else str(value).strip()
                )

                cell.font = Font(
                    name="Arial",
                    size=9,
                )

                cell.fill = PatternFill(
                    "solid",
                    fgColor=row_fill,
                )

                cell.border = thin_border

                if column_number in {
                    2,
                    3,
                    6,
                }:
                    cell.alignment = centered
                else:
                    cell.alignment = left_aligned

            worksheet.row_dimensions[
                row_number
            ].height = 24

        # ---------------------------------------------------------
        # Filtro de tabla
        # ---------------------------------------------------------

        last_row = max(
            header_row,
            header_row + len(contacts),
        )

        worksheet.auto_filter.ref = (
            f"A{header_row}:H{last_row}"
        )

        # Mantener visible la primera fila/columna normal.
        worksheet.sheet_view.topLeftCell = "A1"

        # Configuración de impresión.
        worksheet.print_area = (
            f"A1:H{last_row}"
        )
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.outlinePr.summaryBelow = True

        # ---------------------------------------------------------
        # Guardar en memoria
        # ---------------------------------------------------------

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output.getvalue()

    def find_contacts_file(
        self,
        folder_id: str,
        filename: str,
    ) -> Optional[Dict]:
        """
        Buscar el archivo manual de contactos por su nombre exacto.

        Ejemplo:
            774_Contactos_2S_2026.xlsx
        """
        files = drive_service.list_files(
            folder_id
        )

        expected_name = self.normalize(
            filename
        )

        matches = [
            file
            for file in files
            if self.normalize(
                file.get("name")
            ) == expected_name
        ]

        if not matches:
            return None

        if len(matches) > 1:
            raise CourseContactsError(
                f"Se encontraron varios archivos llamados "
                f"'{filename}' dentro de 2_Contactos. "
                "Debe existir solamente uno."
            )

        contacts_file = matches[0]

        mime_type = contacts_file.get(
            "mimeType",
            "",
        )

        if mime_type != XLSX_MIME:
            raise CourseContactsError(
                f"El archivo '{filename}' existe, "
                "pero no es un archivo XLSX. "
                f"MIME encontrado: {mime_type}"
            )

        return contacts_file

    def source_status(
        self,
        semester: str,
        year: int,
    ) -> Dict:
        """
        Verificar acceso al archivo maestro.
        """
        source = self.load_source(
            semester,
            year,
        )

        docentes = sum(
            1
            for contact in source.records
            if self.normalize(contact.role)
            == "docente"
        )

        auxiliares = sum(
            1
            for contact in source.records
            if self.normalize(contact.role)
            == "auxiliar"
        )

        return {
            "success": True,
            "spreadsheet_id": self.spreadsheet_id(),
            "spreadsheet_name": source.spreadsheet_name,
            "docentes_sheet": source.docentes_sheet,
            "auxiliares_sheet": source.auxiliares_sheet,
            "docentes_count": docentes,
            "auxiliares_count": auxiliares,
            "total_contacts": len(source.records),
            "warnings": source.warnings,
        }

    def preview(
        self,
        courses: Sequence[CourseCatalog],
        semester: str,
        year: int,
    ) -> Dict:
        """
        Mostrar los contactos encontrados y la ubicación exacta
        donde debe colocarse manualmente el Excel de cada curso.
        """
        if not courses:
            raise CourseContactsError(
                "No se recibieron cursos para la vista previa"
            )

        source = self.load_source(
            semester,
            year,
        )

        period = self.period_data(
            semester,
            year,
        )

        period_folder = self.find_period_folder(
            semester,
            year,
        )

        course_results = []

        # Evita buscar varias veces una misma carpeta de área.
        area_folders: Dict[str, Dict] = {}

        for course in courses:
            canonical_area = self.canonical_area(
                course.area
            )

            area_folder_name = self.area_folder_name(
                canonical_area
            )

            contacts = self.contacts_for_course(
                source.records,
                course,
            )

            docentes_count = sum(
                1
                for contact in contacts
                if self.normalize(
                    contact.role
                ) == "docente"
            )

            auxiliares_count = sum(
                1
                for contact in contacts
                if self.normalize(
                    contact.role
                ) == "auxiliar"
            )

            filename = self.contacts_filename(
                course=course,
                semester=semester,
                year=year,
            )

            preview_item = {
                "code": course.code,
                "name": course.name,
                "area": canonical_area,
                "contacts_count": len(contacts),
                "docentes_count": docentes_count,
                "auxiliares_count": auxiliares_count,

                # Nombre listo para copiar.
                "filename": filename,

                # Se completan cuando las carpetas existen.
                "folder_found": False,
                "folder_id": None,
                "folder_link": None,
                "path": (
                    f"{period['folder']}/"
                    f"{area_folder_name}/"
                    f"{course.code}_.../"
                    "2_Contactos"
                ),

                "file_exists": False,
                "file_id": None,
                "file_link": None,
                "status": "pending_folder_lookup",
                "message": "",
            }

            # Buscar la carpeta del área.
            if canonical_area not in area_folders:
                try:
                    area_folders[canonical_area] = (
                        self.find_area_folder(
                            period_folder["id"],
                            canonical_area,
                        )
                    )
                except CourseContactsError as exception:
                    preview_item["status"] = (
                        "area_folder_not_found"
                    )
                    preview_item["message"] = str(
                        exception
                    )
                    course_results.append(
                        preview_item
                    )
                    continue

            area_folder = area_folders[
                canonical_area
            ]

            # Buscar la carpeta del curso.
            course_folder = self.find_course_folder(
                area_folder["id"],
                course,
                semester,
                year,
            )

            if not course_folder:
                preview_item["status"] = (
                    "course_folder_not_found"
                )
                preview_item["path"] = (
                    f"{period['folder']}/"
                    f"{area_folder['name']}/"
                    f"{course.code}_..."
                )
                preview_item["message"] = (
                    "No se encontró la carpeta "
                    f"del curso {course.code}"
                )
                course_results.append(
                    preview_item
                )
                continue

            # Buscar 2_Contactos.
            contacts_folder = self.find_contacts_folder(
                course_folder["id"]
            )

            preview_item["path"] = (
                f"{period['folder']}/"
                f"{area_folder['name']}/"
                f"{course_folder['name']}/"
                "2_Contactos"
            )

            if not contacts_folder:
                preview_item["status"] = (
                    "contacts_folder_not_found"
                )
                preview_item["message"] = (
                    "No se encontró la carpeta "
                    "2_Contactos dentro del curso"
                )
                course_results.append(
                    preview_item
                )
                continue

            folder_link = (
                contacts_folder.get(
                    "webViewLink"
                )
                or (
                    "https://drive.google.com/drive/folders/"
                    f"{contacts_folder['id']}"
                )
            )

            preview_item.update(
                {
                    "folder_found": True,
                    "folder_id": contacts_folder["id"],
                    "folder_link": folder_link,
                    "status": "file_required",
                    "message": (
                        f"Crea o sube '{filename}' "
                        "en esta carpeta."
                    ),
                }
            )

            # Revisar si el archivo ya fue creado manualmente.
            try:
                existing_file = self.find_contacts_file(
                    folder_id=contacts_folder["id"],
                    filename=filename,
                )
            except CourseContactsError as exception:
                preview_item["status"] = (
                    "invalid_contacts_file"
                )
                preview_item["message"] = str(
                    exception
                )
                course_results.append(
                    preview_item
                )
                continue

            if existing_file:
                preview_item.update(
                    {
                        "file_exists": True,
                        "file_id": existing_file["id"],
                        "file_link": (
                            existing_file.get(
                                "webViewLink"
                            )
                        ),
                        "status": "ready",
                        "message": (
                            "El archivo ya existe y "
                            "está listo para actualizarse."
                        ),
                    }
                )

            course_results.append(
                preview_item
            )

        return {
            "success": True,
            "semester": period["code"],
            "year": year,
            "period_folder": period["folder"],
            "courses": course_results,
            "summary": {
                "courses_count": len(course_results),
                "contacts_count": sum(
                    item["contacts_count"]
                    for item in course_results
                ),
                "ready_files_count": sum(
                    1
                    for item in course_results
                    if item["file_exists"]
                ),
                "missing_files_count": sum(
                    1
                    for item in course_results
                    if (
                        item["folder_found"]
                        and not item["file_exists"]
                    )
                ),
            },
            "warnings": source.warnings,
        }

    def create_files(
        self,
        courses: Sequence[CourseCatalog],
        semester: str,
        year: int,
    ) -> Dict:
        """
        Crear o actualizar los Excel de contactos siguiendo esta ruta:

        GOOGLE_DRIVE_STRUCTURE_FOLDER_ID
        / período
        / área
        / curso
        / 2_Contactos
        """
        if not courses:
            raise CourseContactsError(
                "No se recibieron cursos para procesar"
            )

        source = self.load_source(
            semester,
            year,
        )

        period = self.period_data(
            semester,
            year,
        )

        # 1. Buscar la carpeta del período.
        period_folder = self.find_period_folder(
            semester,
            year,
        )

        updated_count = 0
        error_count = 0
        contacts_count = 0
        results = []

        # Caché para no buscar repetidamente una misma carpeta de área.
        area_folders: Dict[str, Dict] = {}

        for course in courses:
            canonical_area = self.canonical_area(
                course.area
            )

            area_folder_name = self.area_folder_name(
                canonical_area
            )

            # 2. Buscar la carpeta del área dentro del período.
            if canonical_area not in area_folders:
                try:
                    area_folders[canonical_area] = (
                        self.find_area_folder(
                            period_folder["id"],
                            canonical_area,
                        )
                    )
                except CourseContactsError as exception:
                    error_count += 1

                    results.append(
                        {
                            "code": course.code,
                            "name": course.name,
                            "area": canonical_area,
                            "success": False,
                            "status": "area_folder_not_found",
                            "contacts_count": 0,
                            "path": (
                                f"{period['folder']}/"
                                f"{area_folder_name}"
                            ),
                            "message": str(exception),
                        }
                    )
                    continue

            area_folder = area_folders[
                canonical_area
            ]

            contacts = self.contacts_for_course(
                source.records,
                course,
            )

            contacts_count += len(contacts)

            # 3. Buscar el curso dentro de la carpeta del área.
            course_folder = self.find_course_folder(
                area_folder["id"],
                course,
                semester,
                year,
            )

            if not course_folder:
                error_count += 1

                results.append(
                    {
                        "code": course.code,
                        "name": course.name,
                        "area": canonical_area,
                        "success": False,
                        "status": "course_folder_not_found",
                        "contacts_count": len(contacts),
                        "path": (
                            f"{period['folder']}/"
                            f"{area_folder['name']}/"
                            f"{course.code}_..."
                        ),
                        "message": (
                            "No se encontró la carpeta del curso "
                            f"{course.code} dentro de "
                            f"'{area_folder['name']}'"
                        ),
                    }
                )
                continue

            # 4. Buscar 2_Contactos dentro del curso.
            contacts_folder = self.find_contacts_folder(
                course_folder["id"]
            )

            full_contacts_path = (
                f"{period['folder']}/"
                f"{area_folder['name']}/"
                f"{course_folder['name']}/"
                "2_Contactos"
            )

            if not contacts_folder:
                error_count += 1

                results.append(
                    {
                        "code": course.code,
                        "name": course.name,
                        "area": canonical_area,
                        "success": False,
                        "status": "contacts_folder_not_found",
                        "contacts_count": len(contacts),
                        "path": full_contacts_path,
                        "message": (
                            "No se encontró la carpeta "
                            "'2_Contactos' dentro del curso"
                        ),
                    }
                )

                continue


            # El archivo se crea manualmente y siempre utiliza este nombre.
            filename = self.contacts_filename(
                course=course,
                semester=semester,
                year=year,
            )

            try:
                existing_file = self.find_contacts_file(
                    folder_id=contacts_folder["id"],
                    filename=filename,
                )
            except CourseContactsError as exception:
                error_count += 1

                results.append(
                    {
                        "code": course.code,
                        "name": course.name,
                        "area": canonical_area,
                        "success": False,
                        "status": "invalid_contacts_file",
                        "contacts_count": len(contacts),
                        "path": full_contacts_path,
                        "folder_id": contacts_folder["id"],
                        "folder_link": (
                            contacts_folder.get("webViewLink")
                            or (
                                "https://drive.google.com/drive/folders/"
                                f"{contacts_folder['id']}"
                            )
                        ),
                        "filename": filename,
                        "message": str(exception),
                    }
                )

                continue

            if not existing_file:
                error_count += 1

                folder_link = (
                    contacts_folder.get("webViewLink")
                    or (
                        "https://drive.google.com/drive/folders/"
                        f"{contacts_folder['id']}"
                    )
                )

                results.append(
                    {
                        "code": course.code,
                        "name": course.name,
                        "area": canonical_area,
                        "success": False,
                        "status": "contacts_file_not_found",
                        "contacts_count": len(contacts),
                        "path": full_contacts_path,
                        "folder_id": contacts_folder["id"],
                        "folder_link": folder_link,
                        "filename": filename,
                        "message": (
                            f"Crea manualmente el archivo '{filename}' "
                            "y colócalo en la carpeta indicada."
                        ),
                    }
                )

                continue

            # Generar el nuevo contenido solamente cuando el archivo exista.
            excel_content = self.build_excel(
                course=course,
                contacts=contacts,
                semester=semester,
                year=year,
            )

            print(
                f"⬆️ Actualizando archivo de contactos | "
                f"curso={course.code} | "
                f"archivo={filename} | "
                f"file_id={existing_file['id']} | "
                f"tamaño={len(excel_content)} bytes"
            )

            updated = drive_service.upload_file(
                file_bytes=excel_content,
                mime_type=XLSX_MIME,
                existing_file_id=existing_file["id"],
            )

            if not updated:
                error_count += 1

                results.append(
                    {
                        "code": course.code,
                        "name": course.name,
                        "area": canonical_area,
                        "success": False,
                        "status": "update_error",
                        "contacts_count": len(contacts),
                        "path": full_contacts_path,
                        "filename": filename,
                        "file_id": existing_file["id"],
                        "message": (
                            f"No se pudo actualizar el archivo "
                            f"'{filename}' en Google Drive"
                        ),
                    }
                )

                continue


            updated_count += 1

            results.append(
                {
                    "code": course.code,
                    "name": course.name,
                    "area": canonical_area,
                    "success": True,
                    "status": "updated",
                    "contacts_count": len(contacts),
                    "path": full_contacts_path,
                    "filename": filename,
                    "file_id": existing_file["id"],
                    "webViewLink": existing_file.get("webViewLink"),
                    "message": (
                        f"El archivo '{filename}' fue actualizado correctamente"
                    ),
                }
            )

        return {
            "success": error_count == 0,
            "message": (
                "Contactos procesados correctamente"
                if error_count == 0
                else "El proceso terminó con errores"
            ),
            "semester": period["code"],
            "year": year,
            "period_folder": period["folder"],
            "root_folder_id": self.root_folder_id(),
            "summary": {
                "courses_processed": len(courses),
                "contacts_processed": contacts_count,
                "updated_count": updated_count,
                "error_count": error_count,
            },
            "results": results,
            "source_warnings": source.warnings,
        }

course_contacts_service = CourseContactsService()