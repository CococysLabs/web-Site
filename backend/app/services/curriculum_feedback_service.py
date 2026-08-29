"""
Servicio de retroalimentación de Diseño Curricular.

Flujo por curso:

1. Localizar:
   periodo/
   área/
   curso/
   3_Planeacion_Curricular/

2. Buscar:
   1_Fortalezas_Debilidades_y_Recomendaciones
   2_Analisis_de_Contexto
   3_Criterios y Expectativas
   4_Analisis_Internacional
   5_Diseño_Curricular
   6_Diseño_Curricular_Retroalimentacion

3. Los documentos 1..4 son opcionales.
4. 5_Diseño_Curricular es obligatorio.
5. 6_Diseño_Curricular_Retroalimentacion solamente es
   obligatorio cuando se desea escribir el resultado.

6. Los Google Sheets 1..5 se exportan a PDF en memoria.
7. Se extrae texto del PDF.
8. Si PDF falla, se usa XLSX en memoria.
9. DeepSeek genera JSON estructurado.
10. Google Sheets API escribe la retroalimentación en el archivo 6.
"""

from __future__ import annotations

import io
import json
import re
import time
import unicodedata
import urllib.error
import urllib.request
import socket

from dataclasses import dataclass
from typing import (
    Any,
    Dict,
    Iterable,
    List,
    Optional,
    Tuple,
)

import openpyxl
from PyPDF2 import PdfReader

from app.config import settings
from app.models.course_catalog import CourseCatalog
from app.services.course_contacts_service import (
    CourseContactsError,
    course_contacts_service,
)
from app.services.drive_service import drive_service
from app.services.google_sheets_service import (
    google_sheets_service,
)


# ============================================================
# CONFIGURACIÓN DEL MÓDULO
# ============================================================

# Se deja en código, NO en .env.
#
# Actualmente corresponde a la raíz usada por la estructura
# curricular del proyecto.
CURRICULUM_ROOT_FOLDER_ID = (
    "1kKtxjCV9cXxkS_BeQv95Ud5M_Q0S77aA"
)

DEEPSEEK_MODEL = "deepseek-v4-flash"

DEEPSEEK_URL = (
    "https://api.deepseek.com/chat/completions"
)

DEEPSEEK_TIMEOUT_SECONDS = 240

DEEPSEEK_MAX_OUTPUT_TOKENS = 16_000

DEEPSEEK_MAX_ATTEMPTS = 3

# No representa tokens.
# Es solamente un límite interno para evitar enviar accidentalmente
# hojas absurdamente grandes generadas por formatos residuales.
MAX_CONTEXT_CHARS = 2_500_000

MAX_XLSX_ROWS = 20_000
MAX_XLSX_COLUMNS = 150
STOP_AFTER_EMPTY_ROWS = 150
MIN_CURRICULUM_CONTENT_CHARS = 2_000

GOOGLE_SHEET_MIME = (
    "application/vnd.google-apps.spreadsheet"
)

PDF_MIME = "application/pdf"

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)


GEMINI_CURRICULUM_MODEL = "gemini-3.5-flash-lite"

GEMINI_STREAM_URL = (
    "https://generativelanguage.googleapis.com/"
    "v1beta/models/{model}:streamGenerateContent?alt=sse"
)


# Este timeout es por inactividad de la conexión.
# Al usar streaming, cada chunk mantiene viva la conexión.
GEMINI_STREAM_TIMEOUT_SECONDS = 120

GEMINI_MAX_OUTPUT_TOKENS = 7000


GROQ_CURRICULUM_MODEL = "openai/gpt-oss-120b"

GROQ_API_URL = (
    "https://api.groq.com/openai/v1/chat/completions"
)

AI_PROVIDER_TIMEOUT_SECONDS = 240

# Groq tiene bastante menos contexto que Gemini.
# Es un límite conservador en CARACTERES, no en tokens.
GROQ_MAX_CONTEXT_CHARS = 350_000

# ============================================================
# ARCHIVOS ESPERADOS
# ============================================================

@dataclass(frozen=True)
class PlanningFileSpec:
    key: str
    name: str
    required_for_analysis: bool = False
    required_for_write: bool = False


PLANNING_FILES: Tuple[
    PlanningFileSpec,
    ...,
] = (
    PlanningFileSpec(
        key="fortalezas_debilidades_recomendaciones",
        name="1_Fortalezas_Debilidades_y_Recomendaciones",
    ),

    PlanningFileSpec(
        key="analisis_contexto",
        name="2_Analisis_de_Contexto",
    ),

    PlanningFileSpec(
        key="criterios_expectativas",
        name="3_Criterios y Expectativas",
    ),

    PlanningFileSpec(
        key="analisis_internacional",
        name="4_Analisis_Internacional",
    ),

    PlanningFileSpec(
        key="diseno_curricular",
        name="5_Diseño_Curricular",
        required_for_analysis=True,
    ),
)


OPTIONAL_CONTEXT_KEYS = {
    "fortalezas_debilidades_recomendaciones",
    "analisis_contexto",
    "criterios_expectativas",
    "analisis_internacional",
}

MAIN_INPUT_KEY = "diseno_curricular"

MATRIX_FOLDER_NAME = (
    "0_Revision_de_Material"
)

MATRIX_FILE_NAME = (
    "02_Matriz observaciones estructura"
)

MATRIX_AI_COLUMN = "G"


CURRICULUM_SHEET_ORDER: Tuple[str, ...] = (
    "competencias",
    "diseno",
    "semana_diagnostico",
    "s2",
    "s3",
    "s4",
    "s5",
    "s6",
    "s7",
    "s8",
    "s9",
    "s10",
    "s11",
    "proyectos",
    "practicas",
    "tareas",
)


CURRICULUM_SHEET_ALIASES: Dict[
    str,
    Tuple[str, ...],
] = {
    "competencias": (
        "Competencias",
    ),

    "diseno": (
        "Diseño",
        "Diseno",
    ),

    "semana_diagnostico": (
        "Semana Diagnostico",
        "Semana Diagnóstico",
        "Semana de Diagnostico",
        "Semana de Diagnóstico",
    ),

    "s2": (
        "S.2",
        "S2",
        "Semana 2",
    ),

    "s3": (
        "S.3",
        "S3",
        "Semana 3",
    ),

    "s4": (
        "S.4",
        "S4",
        "Semana 4",
    ),

    "s5": (
        "S.5",
        "S5",
        "Semana 5",
    ),

    "s6": (
        "S.6",
        "S6",
        "Semana 6",
    ),

    "s7": (
        "S.7",
        "S7",
        "Semana 7",
    ),

    "s8": (
        "S.8",
        "S8",
        "Semana 8",
    ),

    "s9": (
        "S.9",
        "S9",
        "Semana 9",
    ),

    "s10": (
        "S.10",
        "S10",
        "Semana 10",
    ),

    "s11": (
        "S.11",
        "S11",
        "Semana 11",
    ),

    "proyectos": (
        "Proyectos",
    ),

    "practicas": (
        "Practicas",
        "Prácticas",
    ),

    "tareas": (
        "Tareas",
    ),
}

# ============================================================
# SALIDA ESPERADA DE DEEPSEEK
# ============================================================




class CurriculumFeedbackError(
    RuntimeError
):
    pass


class CurriculumFeedbackService:

    def _canonical_curriculum_sheet(
    self,
    sheet_title: str,
    ) -> Optional[str]:

        normalized_title = self.normalize(
            sheet_title
        )

        for key, aliases in (
            CURRICULUM_SHEET_ALIASES.items()
        ):

            for alias in aliases:

                if (
                    normalized_title
                    == self.normalize(alias)
                ):
                    return key

        return None

    def _curriculum_google_sheet_to_text(
        self,
        spreadsheet_id: str,
    ) -> Tuple[
        str,
        Dict[str, Any],
    ]:
        """
        Leer directamente 5_Diseño_Curricular mediante
        Google Sheets API.

        A diferencia de openpyxl data_only=True,
        obtiene los valores visibles/calculados de
        las celdas de Google Sheets.
        """

        sheet_titles = (
            google_sheets_service
            .get_sheet_titles(
                spreadsheet_id
            )
        )

        found_by_key: Dict[
            str,
            str,
        ] = {}

        for title in sheet_titles:

            canonical_key = (
                self._canonical_curriculum_sheet(
                    title
                )
            )

            if canonical_key:

                if canonical_key in found_by_key:

                    raise CurriculumFeedbackError(
                        "5_Diseño_Curricular contiene "
                        "más de una hoja equivalente a "
                        f"'{canonical_key}'"
                    )

                found_by_key[
                    canonical_key
                ] = title

        parts: List[str] = []
        found: List[
            Dict[str, Any]
        ] = []
        missing: List[str] = []

        for key in CURRICULUM_SHEET_ORDER:

            title = found_by_key.get(
                key
            )

            if not title:

                missing.append(
                    key
                )

                parts.append(
                    "========================================\n"
                    f"HOJA NO ENCONTRADA: {key}\n"
                    "========================================"
                )

                continue

            values = (
                google_sheets_service
                .get_sheet_values(
                    spreadsheet_id,
                    title,
                )
            )

            lines: List[str] = []

            non_empty_cells = 0

            for row_number, row in enumerate(
                values,
                start=1,
            ):

                cells: List[str] = []

                for column_index, value in enumerate(
                    row,
                    start=1,
                ):

                    if value is None:
                        continue

                    text = str(
                        value
                    ).strip()

                    if not text:
                        continue

                    text = re.sub(
                        r"\s*\n\s*",
                        " / ",
                        text,
                    )

                    text = re.sub(
                        r"\s+",
                        " ",
                        text,
                    ).strip()

                    column_letter = (
                        openpyxl.utils
                        .get_column_letter(
                            column_index
                        )
                    )

                    cells.append(
                        f"{column_letter}"
                        f"{row_number}={text}"
                    )

                    non_empty_cells += 1

                if cells:

                    lines.append(
                        " | ".join(
                            cells
                        )
                    )

            text = "\n".join(
                lines
            )

            found.append(
                {
                    "key": key,
                    "title": title,
                    "rows": len(
                        values
                    ),
                    "non_empty_cells": (
                        non_empty_cells
                    ),
                    "chars": len(
                        text
                    ),
                }
            )

            parts.append(
                "========================================\n"
                f"HOJA: {title}\n"
                f"CLAVE: {key}\n"
                f"FILAS_LEIDAS: {len(values)}\n"
                f"CELDAS_CON_CONTENIDO: {non_empty_cells}\n"
                "========================================\n"
                f"{text}"
            )

        if not found:

            raise CurriculumFeedbackError(
                "5_Diseño_Curricular no contiene "
                "ninguna de las hojas esperadas"
            )

        return (
            "\n\n".join(
                parts
            ),
            {
                "found": found,
                "missing": missing,
            },
        )

    def _curriculum_xlsx_to_text(
        self,
        content: bytes,
    ) -> Tuple[
        str,
        Dict[str, Any],
    ]:
        """
        Fallback para leer 5_Diseño_Curricular como XLSX.

        Se utiliza solamente si la lectura directa mediante
        Google Sheets API falla.

        data_only=False conserva fórmulas en lugar de convertirlas
        potencialmente en None.
        """

        workbook = (
            openpyxl.load_workbook(
                io.BytesIO(
                    content
                ),
                read_only=True,
                data_only=False,
                keep_links=True,
            )
        )

        found: List[
            Dict[str, Any]
        ] = []

        extracted: Dict[
            str,
            Dict[str, Any]
        ] = {}

        missing: List[str] = []

        try:

            for worksheet in workbook.worksheets:

                canonical_key = (
                    self._canonical_curriculum_sheet(
                        worksheet.title
                    )
                )

                if not canonical_key:
                    continue

                lines: List[str] = []

                non_empty_cells = 0

                maximum_rows = min(
                    worksheet.max_row or 1,
                    MAX_XLSX_ROWS,
                )

                maximum_columns = min(
                    worksheet.max_column or 1,
                    MAX_XLSX_COLUMNS,
                )

                empty_rows = 0

                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=maximum_rows,
                    min_col=1,
                    max_col=maximum_columns,
                ):

                    cells: List[str] = []

                    for cell in row:

                        value = cell.value

                        if value is None:
                            continue

                        cell_text = str(
                            value
                        ).strip()

                        if not cell_text:
                            continue

                        cell_text = re.sub(
                            r"\s*\n\s*",
                            " / ",
                            cell_text,
                        )

                        cell_text = re.sub(
                            r"\s+",
                            " ",
                            cell_text,
                        ).strip()

                        cells.append(
                            f"{cell.coordinate}={cell_text}"
                        )

                        non_empty_cells += 1

                    if not cells:

                        empty_rows += 1

                        if (
                            empty_rows
                            >= STOP_AFTER_EMPTY_ROWS
                        ):
                            break

                        continue

                    empty_rows = 0

                    lines.append(
                        " | ".join(
                            cells
                        )
                    )

                sheet_text = "\n".join(
                    lines
                )

                extracted[
                    canonical_key
                ] = {
                    "title": worksheet.title,
                    "text": sheet_text,
                    "rows": maximum_rows,
                    "non_empty_cells": (
                        non_empty_cells
                    ),
                    "chars": len(
                        sheet_text
                    ),
                }

        finally:

            workbook.close()

        parts: List[str] = []

        for key in CURRICULUM_SHEET_ORDER:

            sheet = extracted.get(
                key
            )

            if not sheet:

                missing.append(
                    key
                )

                parts.append(
                    "========================================\n"
                    f"HOJA NO ENCONTRADA: {key}\n"
                    "========================================"
                )

                continue

            found.append(
                {
                    "key": key,
                    "title": sheet[
                        "title"
                    ],
                    "rows": sheet[
                        "rows"
                    ],
                    "non_empty_cells": sheet[
                        "non_empty_cells"
                    ],
                    "chars": sheet[
                        "chars"
                    ],
                }
            )

            parts.append(
                "========================================\n"
                f"HOJA: {sheet['title']}\n"
                f"CLAVE: {key}\n"
                f"FILAS_LEIDAS: {sheet['rows']}\n"
                "CELDAS_CON_CONTENIDO: "
                f"{sheet['non_empty_cells']}\n"
                "========================================\n"
                f"{sheet['text']}"
            )

        if not found:

            raise CurriculumFeedbackError(
                "5_Diseño_Curricular no contiene "
                "ninguna de las hojas esperadas"
            )

        return (
            "\n\n".join(
                parts
            ),
            {
                "found": found,
                "missing": missing,
            },
        )

    def _validate_curriculum_extraction(
        self,
        text: str,
        workbook_info: Dict[
            str,
            Any
        ],
    ) -> None:
        """
        Impedir que se mande a IA una extracción incompleta
        de 5_Diseño_Curricular.
        """

        if len(
            text.strip()
        ) < MIN_CURRICULUM_CONTENT_CHARS:

            raise CurriculumFeedbackError(
                "5_Diseño_Curricular produjo solamente "
                f"{len(text)} caracteres de contenido. "
                "La extracción parece incompleta."
            )

        required_content_sheets = {
            "competencias",
            "diseno",
            "semana_diagnostico",
            "s2",
            "s3",
            "s4",
            "s5",
            "s6",
            "s7",
            "s8",
            "s9",
            "s10",
            "s11",
        }

        found_sheets = (
            workbook_info.get(
                "found",
                [],
            )
        )

        suspicious_empty = [
            sheet.get(
                "title",
                sheet.get(
                    "key",
                    "desconocida",
                ),
            )

            for sheet in found_sheets

            if (
                sheet.get(
                    "key"
                )
                in required_content_sheets

                and int(
                    sheet.get(
                        "non_empty_cells",
                        0,
                    )
                    or 0
                )
                == 0
            )
        ]

        if suspicious_empty:

            raise CurriculumFeedbackError(
                "La lectura de "
                "5_Diseño_Curricular devolvió "
                "hojas obligatorias sin contenido: "
                + ", ".join(
                    suspicious_empty
                )
            )

    def _extract_curriculum_source(
        self,
        file: Dict[str, Any],
    ) -> Tuple[
        str,
        str,
        List[str],
        Dict[str, Any],
    ]:

        file_id = file[
            "id"
        ]

        file_name = file.get(
            "name",
            file_id,
        )

        mime_type = file.get(
            "mimeType",
            "",
        )

        warnings: List[str] = []

        direct_error: Optional[
            Exception
        ] = None

        # ========================================================
        # 1. GOOGLE SHEETS NATIVO
        # ========================================================

        if mime_type == GOOGLE_SHEET_MIME:

            print(
                "📊 [CURRICULUM] Leyendo "
                "5_Diseño_Curricular directamente "
                "con Google Sheets API | "
                f"archivo={file_name}",
                flush=True,
            )

            try:

                (
                    text,
                    workbook_info,
                ) = (
                    self
                    ._curriculum_google_sheet_to_text(
                        file_id
                    )
                )

                # -----------------------------------------------
                # VALIDAR ANTES DEL RETURN
                # -----------------------------------------------

                self._validate_curriculum_extraction(
                    text,
                    workbook_info,
                )

                if workbook_info.get(
                    "missing"
                ):

                    warnings.append(
                        "5_Diseño_Curricular no contiene "
                        "las siguientes hojas esperadas: "
                        + ", ".join(
                            workbook_info[
                                "missing"
                            ]
                        )
                    )

                print(
                    "✅ [CURRICULUM] "
                    "5_Diseño_Curricular leído "
                    "con Google Sheets API | "
                    f"hojas="
                    f"{len(workbook_info['found'])} | "
                    f"chars={len(text)}",
                    flush=True,
                )

                for sheet in workbook_info.get(
                    "found",
                    [],
                ):

                    print(
                        "   📄 "
                        f"{sheet['title']} | "
                        f"filas={sheet['rows']} | "
                        f"celdas="
                        f"{sheet['non_empty_cells']} | "
                        f"chars={sheet['chars']}",
                        flush=True,
                    )

                return (
                    text,
                    "google_sheets_values",
                    warnings,
                    workbook_info,
                )

            except Exception as exc:

                direct_error = exc

                warnings.append(
                    "No se pudo leer "
                    "5_Diseño_Curricular directamente "
                    "con Google Sheets API. "
                    "Se utilizará XLSX como respaldo. "
                    f"Motivo: {exc}"
                )

                print(
                    "⚠️ [CURRICULUM] "
                    "Sheets API falló; "
                    "usando XLSX fallback | "
                    f"tipo={type(exc).__name__} | "
                    f"error={exc}",
                    flush=True,
                )

        # ========================================================
        # 2. COMPROBAR FORMATO PARA FALLBACK
        # ========================================================

        if mime_type not in {
            GOOGLE_SHEET_MIME,
            XLSX_MIME,
            "application/vnd.ms-excel",
        }:

            raise CurriculumFeedbackError(
                f"'{file_name}' debe ser un "
                "Google Sheets o XLSX"
            )

        # ========================================================
        # 3. XLSX FALLBACK
        # ========================================================

        print(
            "📥 [CURRICULUM] "
            "Leyendo 5_Diseño_Curricular "
            "mediante XLSX fallback | "
            f"archivo={file_name}",
            flush=True,
        )

        content = (
            drive_service.download_file(
                file_id
            )
        )

        if not content:

            raise CurriculumFeedbackError(
                "No se pudo obtener "
                "5_Diseño_Curricular como XLSX"
            )

        try:

            (
                text,
                workbook_info,
            ) = (
                self._curriculum_xlsx_to_text(
                    content
                )
            )

        except Exception as exc:

            message = (
                "No se pudo leer "
                "5_Diseño_Curricular mediante XLSX"
            )

            if direct_error:

                message += (
                    ". La lectura directa con "
                    "Google Sheets API también falló: "
                    f"{direct_error}"
                )

            message += (
                f". Error XLSX: {exc}"
            )

            raise CurriculumFeedbackError(
                message
            ) from exc

        finally:

            content = None

        # ========================================================
        # 4. VALIDAR FALLBACK
        # ========================================================

        try:

            self._validate_curriculum_extraction(
                text,
                workbook_info,
            )

        except CurriculumFeedbackError as exc:

            if direct_error:

                raise CurriculumFeedbackError(
                    "La lectura directa de "
                    "5_Diseño_Curricular falló y el "
                    "XLSX de respaldo también produjo "
                    "una extracción incompleta. "
                    f"Sheets API: {direct_error}. "
                    f"XLSX: {exc}"
                ) from exc

            raise

        if workbook_info.get(
            "missing"
        ):

            warnings.append(
                "5_Diseño_Curricular no contiene "
                "las siguientes hojas esperadas: "
                + ", ".join(
                    workbook_info[
                        "missing"
                    ]
                )
            )

        print(
            "✅ [CURRICULUM] "
            "5_Diseño_Curricular leído mediante "
            "XLSX fallback | "
            f"hojas="
            f"{len(workbook_info['found'])} | "
            f"chars={len(text)}",
            flush=True,
        )

        for sheet in workbook_info.get(
            "found",
            [],
        ):

            print(
                "   📄 "
                f"{sheet['title']} | "
                f"filas={sheet['rows']} | "
                f"celdas="
                f"{sheet['non_empty_cells']} | "
                f"chars={sheet['chars']}",
                flush=True,
            )

        return (
            text,
            "xlsx_structured_fallback",
            warnings,
            workbook_info,
        )
    
    def _call_ai(
        self,
        system_message: str,
        user_message: str,
        response_schema: Optional[
            Dict[str, Any]
        ] = None,
    ) -> Dict[str, Any]:
        """
        Cadena de proveedores para retroalimentación curricular.

        1. Gemini 2.5 Flash-Lite
        2. DeepSeek, si existe key y tiene saldo
        3. Groq, únicamente si el contexto cabe
        """

        errors: List[str] = []

        # ========================================================
        # 1. GEMINI
        # ========================================================

        try:

            return self._call_gemini(
                system_message,
                user_message,
                response_schema=(
                    response_schema
                ),
            )

        except Exception as exc:

            message = str(exc)

            errors.append(
                f"Gemini: {message}"
            )

            print(
                "⚠️ [CURRICULUM] Gemini no disponible | "
                f"tipo={type(exc).__name__} | "
                f"mensaje={message}",
                flush=True,
            )

        # ========================================================
        # 2. DEEPSEEK
        # ========================================================

        deepseek_key = str(
            getattr(
                settings,
                "DEEPSEEK_API_KEY",
                "",
            )
            or ""
        ).strip()

        if deepseek_key:

            try:

                result = self._call_deepseek(
                    system_message,
                    user_message,
                )

                # Nuestro método anterior no incluía provider.
                result[
                    "provider"
                ] = "deepseek"

                return result

            except Exception as exc:

                message = str(exc)

                errors.append(
                    f"DeepSeek: {message}"
                )

                if (
                    "HTTP 402"
                    in message
                    or "Insufficient Balance"
                    in message
                ):

                    print(
                        "⚠️ [CURRICULUM] DeepSeek sin saldo. "
                        "Pasando directamente a Groq.",
                        flush=True,
                    )

                else:

                    print(
                        "⚠️ [CURRICULUM] DeepSeek falló | "
                        f"{message}",
                        flush=True,
                    )

        # ========================================================
        # 3. GROQ
        # ========================================================

        try:

            return self._call_groq(
                system_message,
                user_message,
            )

        except Exception as exc:

            message = str(exc)

            errors.append(
                f"Groq: {message}"
            )

            print(
                "⚠️ [CURRICULUM] Groq no disponible | "
                f"{message}",
                flush=True,
            )

        # ========================================================
        # NINGUNO FUNCIONÓ
        # ========================================================

        raise CurriculumFeedbackError(
            "Ningún proveedor de IA pudo completar "
            "la retroalimentación. "
            + " | ".join(errors)
        )

    def _call_groq(
        self,
        system_message: str,
        user_message: str,
    ) -> Dict[str, Any]:

        api_key = str(
            getattr(
                settings,
                "GROQ_API_KEY",
                "",
            )
            or ""
        ).strip()

        if not api_key:
            raise CurriculumFeedbackError(
                "GROQ_API_KEY no está configurada"
            )

        total_chars = (
            len(system_message)
            + len(user_message)
        )

        if (
            total_chars
            > GROQ_MAX_CONTEXT_CHARS
        ):
            raise CurriculumFeedbackError(
                "El contexto es demasiado grande "
                "para utilizar Groq como respaldo "
                f"({total_chars} caracteres)"
            )

        payload = json.dumps(
            {
                "model": GROQ_CURRICULUM_MODEL,

                "messages": [
                    {
                        "role": "system",
                        "content": system_message,
                    },
                    {
                        "role": "user",
                        "content": user_message,
                    },
                ],

                "temperature": 0.1,

                "max_tokens": 12000,

                "response_format": {
                    "type": "json_object",
                },
            },
            ensure_ascii=False,
        ).encode(
            "utf-8"
        )

        print(
            "🤖 [CURRICULUM] Groq | "
            f"modelo={GROQ_CURRICULUM_MODEL} | "
            f"request_bytes={len(payload)}",
            flush=True,
        )

        request = urllib.request.Request(
            GROQ_API_URL,
            data=payload,
            headers={
                "Authorization": (
                    f"Bearer {api_key}"
                ),
                "Content-Type": (
                    "application/json"
                ),
            },
            method="POST",
        )

        try:

            with urllib.request.urlopen(
                request,
                timeout=AI_PROVIDER_TIMEOUT_SECONDS,
            ) as response:

                response_json = json.loads(
                    response.read().decode(
                        "utf-8"
                    )
                )

            choices = response_json.get(
                "choices",
                [],
            )

            if not choices:
                raise CurriculumFeedbackError(
                    "Groq no devolvió ninguna respuesta"
                )

            raw = (
                choices[0]
                .get(
                    "message",
                    {}
                )
                .get(
                    "content",
                    "",
                )
            )

            parsed = self._clean_json_response(
                raw
            )

            usage = response_json.get(
                "usage",
                {},
            )

            print(
                "✅ [CURRICULUM] Groq respondió | "
                f"modelo={GROQ_CURRICULUM_MODEL}",
                flush=True,
            )

            return {
                "data": parsed,

                "provider": "groq",

                "model": (
                    GROQ_CURRICULUM_MODEL
                ),

                "usage": usage,
            }

        except urllib.error.HTTPError as exc:

            try:
                body = (
                    exc.read()
                    .decode(
                        "utf-8",
                        errors="replace",
                    )
                )
            except Exception:
                body = ""

            raise CurriculumFeedbackError(
                f"Groq HTTP {exc.code}: "
                f"{body[:700]}"
            ) from exc

        except (
            urllib.error.URLError,
            TimeoutError,
        ) as exc:

            raise CurriculumFeedbackError(
                f"Groq no respondió: {exc}"
            ) from exc

    def _call_gemini(
        self,
        system_message: str,
        user_message: str,
        response_schema: Optional[
            Dict[str, Any]
        ] = None,
    ) -> Dict[str, Any]:
        """
        Ejecuta Gemini mediante streamGenerateContent.

        IMPORTANTE:
        No esperamos a que Google cierre físicamente la conexión SSE.
        En cuanto Gemini devuelve finishReason, finalizamos nosotros
        el stream para evitar que la petición HTTP quede colgada.
        """

        keys = self._get_gemini_keys()

        if not keys:
            raise CurriculumFeedbackError(
                "No hay ninguna GEMINI_API_KEY configurada"
            )

        generation_config = {
            "temperature": 0.1,

            "maxOutputTokens": (
                GEMINI_MAX_OUTPUT_TOKENS
            ),

            "responseMimeType": (
                "application/json"
            ),
        }

        if response_schema:

            generation_config[
                "responseSchema"
            ] = response_schema

        payload = json.dumps(
            {
                "systemInstruction": {
                    "parts": [
                        {
                            "text": system_message,
                        }
                    ],
                },

                "contents": [
                    {
                        "role": "user",
                        "parts": [
                            {
                                "text": user_message,
                            }
                        ],
                    }
                ],

                "generationConfig": (
                    generation_config
                ),
            },
            ensure_ascii=False,
        ).encode("utf-8")

        last_error: Optional[Exception] = None

        for key_index, api_key in enumerate(
            keys,
            start=1,
        ):

            url = GEMINI_STREAM_URL.format(
                model=GEMINI_CURRICULUM_MODEL
            )

            print(
                "\n"
                "🤖 [CURRICULUM] Gemini STREAM | "
                f"modelo={GEMINI_CURRICULUM_MODEL} | "
                f"key={key_index}/{len(keys)} | "
                f"request_bytes={len(payload)}",
                flush=True,
            )

            request = urllib.request.Request(
                url,
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-goog-api-key": api_key,
                    "Accept": "text/event-stream",

                    # No queremos depender de una conexión
                    # keep-alive después de terminar la generación.
                    "Connection": "close",
                },
                method="POST",
            )

            started_at = time.perf_counter()

            text_parts: List[str] = []
            usage_metadata: Dict[str, Any] = {}

            finish_reason = None
            chunk_count = 0

            try:

                with urllib.request.urlopen(
                    request,
                    timeout=GEMINI_STREAM_TIMEOUT_SECONDS,
                ) as response:

                    print(
                        "📡 [CURRICULUM] Gemini conexión establecida",
                        flush=True,
                    )

                    for raw_line in response:

                        if not raw_line:
                            continue

                        line = (
                            raw_line
                            .decode(
                                "utf-8",
                                errors="replace",
                            )
                            .strip()
                        )

                        if not line.startswith("data:"):
                            continue

                        json_text = (
                            line[len("data:"):]
                            .strip()
                        )

                        if not json_text:
                            continue

                        if json_text == "[DONE]":
                            print(
                                "🏁 [CURRICULUM] Gemini envió [DONE]",
                                flush=True,
                            )
                            break

                        try:
                            event = json.loads(
                                json_text
                            )

                        except json.JSONDecodeError:

                            print(
                                "⚠️ [CURRICULUM] "
                                "Chunk SSE inválido ignorado",
                                flush=True,
                            )
                            continue

                        chunk_count += 1

                        # ==========================================
                        # USO / TOKENS
                        # ==========================================

                        if event.get("usageMetadata"):
                            usage_metadata = event[
                                "usageMetadata"
                            ]

                        # ==========================================
                        # CANDIDATO
                        # ==========================================

                        candidates = event.get(
                            "candidates",
                            [],
                        )

                        if candidates:

                            candidate = candidates[0]

                            content = candidate.get(
                                "content",
                                {},
                            )

                            parts = content.get(
                                "parts",
                                [],
                            )

                            # Guardar primero TODO el texto
                            # del evento actual.
                            for part in parts:

                                text = part.get(
                                    "text"
                                )

                                if text:
                                    text_parts.append(
                                        text
                                    )

                            candidate_finish_reason = (
                                candidate.get(
                                    "finishReason"
                                )
                            )

                            if candidate_finish_reason:

                                finish_reason = (
                                    candidate_finish_reason
                                )

                        # ==========================================
                        # LOG
                        # ==========================================

                        if (
                            chunk_count == 1
                            or chunk_count % 10 == 0
                            or finish_reason
                        ):

                            current_chars = sum(
                                len(item)
                                for item in text_parts
                            )

                            elapsed = (
                                time.perf_counter()
                                - started_at
                            )

                            print(
                                "📥 [CURRICULUM] Gemini generando | "
                                f"chunks={chunk_count} | "
                                f"chars={current_chars} | "
                                f"finish_reason={finish_reason} | "
                                f"segundos={elapsed:.1f}",
                                flush=True,
                            )

                        # ==========================================
                        # CORRECCIÓN IMPORTANTE
                        # ==========================================
                        #
                        # NO esperamos a que Google cierre
                        # físicamente la conexión HTTP.
                        #
                        # STOP significa que el modelo ya terminó.
                        #
                        # También salimos para MAX_TOKENS y otros
                        # finishReason; se validarán abajo.
                        # ==========================================

                        if finish_reason:
                            print(
                                "🛑 [CURRICULUM] "
                                "Fin lógico del stream detectado | "
                                f"finish_reason={finish_reason}",
                                flush=True,
                            )
                            break

                # Al salir del with urllib cierra el socket.
                elapsed = (
                    time.perf_counter()
                    - started_at
                )

                raw = "".join(
                    text_parts
                ).strip()

                print(
                    "✅ [CURRICULUM] Gemini stream terminado | "
                    f"chunks={chunk_count} | "
                    f"chars={len(raw)} | "
                    f"finish_reason={finish_reason} | "
                    f"segundos={elapsed:.2f}",
                    flush=True,
                )

                # ==============================================
                # VALIDACIONES
                # ==============================================

                if not raw:

                    raise CurriculumFeedbackError(
                        "Gemini terminó el stream "
                        "sin devolver contenido"
                    )

                if finish_reason == "MAX_TOKENS":

                    raise CurriculumFeedbackError(
                        "Gemini alcanzó el límite máximo "
                        "de tokens de salida"
                    )

                if finish_reason in {
                    "MALFORMED_FUNCTION_CALL",
                    "SAFETY",
                    "RECITATION",
                    "BLOCKLIST",
                    "PROHIBITED_CONTENT",
                    "SPII",
                }:

                    raise CurriculumFeedbackError(
                        "Gemini detuvo la generación. "
                        f"finish_reason={finish_reason}"
                    )

                # STOP es el resultado normal.
                if (
                    finish_reason
                    and finish_reason != "STOP"
                ):

                    print(
                        "⚠️ [CURRICULUM] Gemini terminó con "
                        f"finish_reason={finish_reason}",
                        flush=True,
                    )

                # ==============================================
                # JSON
                # ==============================================

                print(
                    "🔎 [CURRICULUM] Parseando JSON de Gemini...",
                    flush=True,
                )

                parsed = (
                    self._clean_json_response(
                        raw
                    )
                )

                print(
                    "✅ [CURRICULUM] JSON de Gemini parseado",
                    flush=True,
                )

                return {
                    "data": parsed,

                    "provider": "gemini",

                    "model": (
                        GEMINI_CURRICULUM_MODEL
                    ),

                    "usage": {
                        "prompt_tokens": (
                            usage_metadata.get(
                                "promptTokenCount"
                            )
                        ),

                        "completion_tokens": (
                            usage_metadata.get(
                                "candidatesTokenCount"
                            )
                        ),

                        "total_tokens": (
                            usage_metadata.get(
                                "totalTokenCount"
                            )
                        ),

                        "cached_tokens": (
                            usage_metadata.get(
                                "cachedContentTokenCount"
                            )
                        ),
                    },
                }

            # ==================================================
            # TIMEOUT
            # ==================================================

            except (
                socket.timeout,
                TimeoutError,
            ) as exc:

                last_error = exc

                elapsed = (
                    time.perf_counter()
                    - started_at
                )

                print(
                    "⏱️ [CURRICULUM] Gemini timeout | "
                    f"key={key_index}/{len(keys)} | "
                    f"chunks_recibidos={chunk_count} | "
                    f"chars_recibidos="
                    f"{sum(len(x) for x in text_parts)} | "
                    f"finish_reason={finish_reason} | "
                    f"segundos={elapsed:.2f}",
                    flush=True,
                )

                continue

            # ==================================================
            # HTTP
            # ==================================================

            except urllib.error.HTTPError as exc:

                try:
                    body = (
                        exc.read()
                        .decode(
                            "utf-8",
                            errors="replace",
                        )
                    )

                except Exception:
                    body = ""

                last_error = (
                    CurriculumFeedbackError(
                        f"Gemini HTTP {exc.code}: "
                        f"{body[:1000]}"
                    )
                )

                print(
                    "⚠️ [CURRICULUM] Gemini HTTP | "
                    f"key={key_index}/{len(keys)} | "
                    f"status={exc.code} | "
                    f"mensaje={body[:300]}",
                    flush=True,
                )

                if exc.code in {
                    408,
                    429,
                    500,
                    502,
                    503,
                    504,
                }:
                    continue

                if exc.code == 400:

                    # Una API key inválida es un problema de ESA key,
                    # no necesariamente de la petición.
                    #
                    # Por lo tanto, probamos la siguiente key.
                    if (
                        "API_KEY_INVALID" in body
                        or "API key not valid" in body
                    ):

                        print(
                            "🔑 [CURRICULUM] "
                            "Gemini key inválida; "
                            "probando siguiente key | "
                            f"key={key_index}/{len(keys)}",
                            flush=True,
                        )

                        continue

                    # Otros HTTP 400 sí suelen indicar que
                    # la petición/schema enviado es inválido.
                    break

                if exc.code in {
                    401,
                    403,
                }:
                    continue

                break

            # ==================================================
            # RED
            # ==================================================

            except urllib.error.URLError as exc:

                last_error = exc

                print(
                    "🌐 [CURRICULUM] Error de red "
                    "con Gemini | "
                    f"key={key_index}/{len(keys)} | "
                    f"error={exc}",
                    flush=True,
                )

                continue

            # ==================================================
            # RESPUESTA INVALIDA
            # ==================================================

            except CurriculumFeedbackError as exc:

                last_error = exc

                print(
                    "⚠️ [CURRICULUM] Respuesta de "
                    "Gemini no válida | "
                    f"key={key_index}/{len(keys)} | "
                    f"error={exc}",
                    flush=True,
                )

                continue

            except Exception as exc:

                last_error = exc

                print(
                    "❌ [CURRICULUM] Error inesperado "
                    "en Gemini | "
                    f"tipo={type(exc).__name__} | "
                    f"error={exc}",
                    flush=True,
                )

                continue

        raise CurriculumFeedbackError(
            "Gemini no pudo completar el análisis "
            f"con ninguna de las {len(keys)} key(s): "
            f"{last_error}"
        )
        

    def _get_gemini_keys(
        self,
    ) -> List[str]:
        keys: List[str] = []

        primary = str(
            getattr(
                settings,
                "GEMINI_API_KEY",
                "",
            )
            or ""
        ).strip()

        if primary and primary not in {
            "kjkj",
            "tu_api_key_de_gemini",
        }:
            keys.append(primary)

        extra_raw = str(
            getattr(
                settings,
                "GEMINI_API_KEYS",
                "",
            )
            or ""
        )

        for item in extra_raw.split(","):
            key = item.strip()

            if (
                key
                and key not in keys
            ):
                keys.append(key)

        return keys

    @staticmethod
    def _clean_json_response(
        raw: str,
    ) -> Dict[str, Any]:
        raw = str(raw or "").strip()

        raw = re.sub(
            r"^```json?\s*",
            "",
            raw,
            flags=re.IGNORECASE,
        )

        raw = re.sub(
            r"\s*```$",
            "",
            raw,
        )

        if not raw:
            raise CurriculumFeedbackError(
                "El proveedor de IA devolvió una respuesta vacía"
            )

        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise CurriculumFeedbackError(
                f"El proveedor devolvió JSON inválido: {exc}"
            ) from exc

        if not isinstance(parsed, dict):
            raise CurriculumFeedbackError(
                "La respuesta de IA no es un objeto JSON"
            )

        return parsed

    # ========================================================
    # NORMALIZACIÓN
    # ========================================================

    @staticmethod
    def normalize(
        value: object,
    ) -> str:
        """
        Comparación tolerante:

        5_Diseño_Curricular
        5 Diseño Curricular
        5-Diseño Curricular
        5_Diseño_Curricular.xlsx

        se consideran equivalentes.
        """
        text = str(
            value or ""
        ).strip()

        text = re.sub(
            r"\.(xlsx|xls|pdf)$",
            "",
            text,
            flags=re.IGNORECASE,
        )

        text = text.replace(
            "_",
            " ",
        )

        text = text.replace(
            "-",
            " ",
        )

        text = re.sub(
            r"\s+",
            " ",
            text,
        ).strip().lower()

        normalized = unicodedata.normalize(
            "NFKD",
            text,
        )

        normalized = "".join(
            character
            for character in normalized
            if not unicodedata.combining(
                character
            )
        )

        return normalized

    # ========================================================
    # UBICACIÓN DE CARPETAS
    # ========================================================

    def _find_period_folder(
        self,
        semester: str,
        year: int,
    ) -> Dict[str, Any]:
        """
        Soporta los dos casos:

        CASO A:
        CURRICULUM_ROOT_FOLDER_ID ya apunta a:
        2026_Segundo_Semestre

        CASO B:
        CURRICULUM_ROOT_FOLDER_ID apunta a una carpeta superior
        que contiene:
        2026_Segundo_Semestre
        """
        period = (
            course_contacts_service
            .period_data(
                semester,
                year,
            )
        )

        expected_name = period[
            "folder"
        ]

        root = (
            drive_service
            .get_file_metadata(
                CURRICULUM_ROOT_FOLDER_ID
            )
        )

        if not root:
            raise CurriculumFeedbackError(
                "No se pudo acceder a la carpeta raíz "
                "de Planeación Curricular"
            )

        root_name = root.get(
            "name",
            "",
        )

        if (
            self.normalize(root_name)
            == self.normalize(
                expected_name
            )
        ):
            return {
                **root,
                "id": (
                    root.get("id")
                    or CURRICULUM_ROOT_FOLDER_ID
                ),
            }

        period_folder = (
            drive_service.find_folder(
                expected_name,
                CURRICULUM_ROOT_FOLDER_ID,
            )
        )

        if not period_folder:
            raise CurriculumFeedbackError(
                "No se encontró la carpeta del período "
                f"'{expected_name}'"
            )

        return period_folder

    def _find_planning_folder(
        self,
        course_folder_id: str,
    ) -> Dict[str, Any]:

        folders = (
            drive_service
            .list_folders(
                course_folder_id
            )
        )

        exact_matches = [
            folder
            for folder in folders
            if self.normalize(
                folder.get(
                    "name"
                )
            )
            == self.normalize(
                "3_Planeacion_Curricular"
            )
        ]

        if len(exact_matches) == 1:
            return exact_matches[0]

        if len(exact_matches) > 1:
            raise CurriculumFeedbackError(
                "Hay más de una carpeta llamada "
                "3_Planeacion_Curricular"
            )

        # Respaldo por si alguien quitó el prefijo 3.
        fallback = [
            folder
            for folder in folders
            if self.normalize(
                folder.get(
                    "name"
                )
            )
            == self.normalize(
                "Planeacion Curricular"
            )
        ]

        if len(fallback) == 1:
            return fallback[0]

        if len(fallback) > 1:
            raise CurriculumFeedbackError(
                "Hay más de una carpeta de "
                "Planeación Curricular"
            )

        raise CurriculumFeedbackError(
            "No se encontró la carpeta "
            "3_Planeacion_Curricular"
        )

    def locate_course(
        self,
        course: CourseCatalog,
        semester: str,
        year: int,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Localizar:

        periodo/
            área/
                curso/
                    3_Planeacion_Curricular/
        """
        try:
            period_folder = (
                self._find_period_folder(
                    semester,
                    year,
                )
            )

            canonical_area = (
                course_contacts_service
                .canonical_area(
                    course.area
                )
            )

            area_folder = (
                course_contacts_service
                .find_area_folder(
                    period_folder["id"],
                    canonical_area,
                )
            )

            course_folder = (
                course_contacts_service
                .find_course_folder(
                    area_folder["id"],
                    course,
                    semester,
                    year,
                )
            )

        except CourseContactsError as exc:
            raise CurriculumFeedbackError(
                str(exc)
            ) from exc

        if not course_folder:
            raise CurriculumFeedbackError(
                "No se encontró de forma inequívoca "
                f"la carpeta del curso {course.code}"
            )

        planning_folder = (
            self._find_planning_folder(
                course_folder["id"]
            )
        )

        return {
            "period": period_folder,
            "area": area_folder,
            "course": course_folder,
            "planning": planning_folder,
        }

    # ========================================================
    # LOCALIZAR ARCHIVOS 1..6
    # ========================================================

    def locate_planning_files(
        self,
        planning_folder_id: str,
    ) -> Dict[
        str,
        Optional[Dict[str, Any]],
    ]:

        files = (
            drive_service
            .list_files(
                planning_folder_id
            )
        )

        result: Dict[
            str,
            Optional[Dict[str, Any]],
        ] = {}

        for spec in PLANNING_FILES:

            expected = self.normalize(
                spec.name
            )

            matches = [
                file
                for file in files
                if self.normalize(
                    file.get(
                        "name"
                    )
                )
                == expected
            ]

            if len(matches) > 1:
                raise CurriculumFeedbackError(
                    "Hay más de un archivo que coincide con "
                    f"'{spec.name}'"
                )

            result[
                spec.key
            ] = (
                matches[0]
                if matches
                else None
            )

        return result

    # ========================================================
    # PREVIEW
    # ========================================================

    def preview_course(
        self,
        course: CourseCatalog,
        semester: str,
        year: int,
    ) -> Dict[str, Any]:
        """
        Comprobar un curso sin utilizar IA ni modificar archivos.

        Valida:

        3_Planeacion_Curricular/
            1_Fortalezas_Debilidades_y_Recomendaciones   opcional
            2_Analisis_de_Contexto                       opcional
            3_Criterios y Expectativas                   opcional
            4_Analisis_Internacional                     opcional
            5_Diseño_Curricular                          obligatorio

        0_Revision_de_Material/
            02_Matriz observaciones estructura           obligatorio
        """

        try:
            # ====================================================
            # 1. LOCALIZAR CURSO
            # ====================================================

            locations = self.locate_course(
                course,
                semester,
                year,
            )

            # ====================================================
            # 2. ARCHIVOS DE PLANEACIÓN
            # ====================================================

            files = self.locate_planning_files(
                locations["planning"]["id"]
            )

            # ====================================================
            # 3. CARPETA 0_Revision_de_Material
            # ====================================================

            revision_folder = (
                self._find_revision_folder(
                    locations["course"]["id"]
                )
            )

            matrix_file = None
            matrix_sheet_title = None
            matrix_targets: List[
                Dict[str, Any]
            ] = []

            # ====================================================
            # 4. MATRIZ
            # ====================================================

            if revision_folder:

                matrix_file = (
                    self._find_matrix_file(
                        revision_folder["id"]
                    )
                )

                if matrix_file:

                    (
                        matrix_sheet_title,
                        matrix_targets,
                    ) = self._matrix_targets(
                        matrix_file["id"]
                    )

        except CurriculumFeedbackError as exc:

            return {
                "code": str(
                    course.code
                ),

                "name": course.name,

                "area": course.area,

                "success": False,

                "ready_for_analysis": False,

                "ready_for_write": False,

                "warnings": [],

                "error": str(
                    exc
                ),

                "locations": {},

                "files": {},
            }

        except Exception as exc:

            return {
                "code": str(
                    course.code
                ),

                "name": course.name,

                "area": course.area,

                "success": False,

                "ready_for_analysis": False,

                "ready_for_write": False,

                "warnings": [],

                "error": (
                    "Error inesperado durante la comprobación: "
                    f"{exc}"
                ),

                "locations": {},

                "files": {},
            }

        # ========================================================
        # 5. ADVERTENCIAS
        # ========================================================

        warnings: List[str] = []

        # 1..4 son opcionales.
        for spec in PLANNING_FILES[:4]:

            if not files.get(
                spec.key
            ):

                warnings.append(
                    f"No se encontró {spec.name}. "
                    "El análisis podrá continuar, "
                    "pero dispondrá de menos contexto."
                )

        # ========================================================
        # 6. 5_Diseño_Curricular
        # ========================================================

        main_file = files.get(
            MAIN_INPUT_KEY
        )

        main_file_valid = False

        if not main_file:

            warnings.append(
                "No se encontró 5_Diseño_Curricular. "
                "El curso no puede analizarse."
            )

        else:

            main_mime = main_file.get(
                "mimeType",
                "",
            )

            main_file_valid = (
                main_mime
                in {
                    GOOGLE_SHEET_MIME,
                    XLSX_MIME,
                    "application/vnd.ms-excel",
                }
            )

            if not main_file_valid:

                warnings.append(
                    "5_Diseño_Curricular existe, "
                    "pero no es un Google Sheets o "
                    "archivo Excel compatible."
                )

        # ========================================================
        # 7. MATRIZ
        # ========================================================

        if not revision_folder:

            warnings.append(
                "No se encontró la carpeta "
                "0_Revision_de_Material."
            )

        elif not matrix_file:

            warnings.append(
                "No se encontró "
                "02_Matriz observaciones estructura."
            )

        elif not matrix_targets:

            warnings.append(
                "02_Matriz observaciones estructura "
                "no contiene secciones reconocibles."
            )

        matrix_valid = bool(
            matrix_file
            and matrix_targets
        )

        # ========================================================
        # 8. ESTADO
        # ========================================================

        ready_for_analysis = bool(
            main_file_valid
            and matrix_valid
        )

        ready_for_write = (
            ready_for_analysis
        )

        # ========================================================
        # 9. LOCATIONS
        # ========================================================

        locations_response = {
            key: {
                "id": value.get(
                    "id"
                ),

                "name": value.get(
                    "name"
                ),

                "webViewLink": value.get(
                    "webViewLink"
                ),
            }

            for key, value
            in locations.items()
        }

        locations_response[
            "revision"
        ] = (
            {
                "id": revision_folder.get(
                    "id"
                ),

                "name": revision_folder.get(
                    "name"
                ),

                "webViewLink": revision_folder.get(
                    "webViewLink"
                ),
            }

            if revision_folder
            else None
        )

        # ========================================================
        # 10. ARCHIVOS 1..5
        # ========================================================

        files_response = {
            spec.key: {
                "expected_name": spec.name,

                "found": bool(
                    files.get(
                        spec.key
                    )
                ),

                "optional": (
                    spec.key
                    in OPTIONAL_CONTEXT_KEYS
                ),

                "id": (
                    files[
                        spec.key
                    ].get(
                        "id"
                    )
                    if files.get(
                        spec.key
                    )
                    else None
                ),

                "name": (
                    files[
                        spec.key
                    ].get(
                        "name"
                    )
                    if files.get(
                        spec.key
                    )
                    else None
                ),

                "mimeType": (
                    files[
                        spec.key
                    ].get(
                        "mimeType"
                    )
                    if files.get(
                        spec.key
                    )
                    else None
                ),

                "webViewLink": (
                    files[
                        spec.key
                    ].get(
                        "webViewLink"
                    )
                    if files.get(
                        spec.key
                    )
                    else None
                ),
            }

            for spec
            in PLANNING_FILES
        }

        # ========================================================
        # 11. MATRIZ
        # ========================================================

        files_response[
            "matrix"
        ] = {
            "expected_name": (
                MATRIX_FILE_NAME
            ),

            "found": bool(
                matrix_file
            ),

            "optional": False,

            "id": (
                matrix_file.get(
                    "id"
                )
                if matrix_file
                else None
            ),

            "name": (
                matrix_file.get(
                    "name"
                )
                if matrix_file
                else None
            ),

            "mimeType": (
                matrix_file.get(
                    "mimeType"
                )
                if matrix_file
                else None
            ),

            "webViewLink": (
                matrix_file.get(
                    "webViewLink"
                )
                if matrix_file
                else None
            ),

            "sheet": (
                matrix_sheet_title
            ),

            "targets": len(
                matrix_targets
            ),

            "sections": (
                matrix_targets
            ),
        }

        # ========================================================
        # 12. RESULTADO
        # ========================================================

        return {
            "code": str(
                course.code
            ),

            "name": course.name,

            "area": (
                course_contacts_service
                .canonical_area(
                    course.area
                )
            ),

            "success": True,

            "ready_for_analysis": (
                ready_for_analysis
            ),

            "ready_for_write": (
                ready_for_write
            ),

            "warnings": warnings,

            "error": None,

            "locations": (
                locations_response
            ),

            "files": (
                files_response
            ),
        }


    def preview(
        self,
        courses: Iterable[
            CourseCatalog
        ],
        semester: str,
        year: int,
    ) -> Dict[str, Any]:
        """
        Comprobar varios cursos sin llamar a IA.
        """

        items = [
            self.preview_course(
                course,
                semester,
                year,
            )
            for course
            in courses
        ]

        return {
            "success": True,

            "semester": (
                course_contacts_service
                .normalize_semester(
                    semester
                )
            ),

            "year": year,

            "summary": {
                "total_courses": len(
                    items
                ),

                "ready_for_analysis": sum(
                    1
                    for item in items
                    if item.get(
                        "ready_for_analysis"
                    )
                ),

                "ready_for_write": sum(
                    1
                    for item in items
                    if item.get(
                        "ready_for_write"
                    )
                ),

                "with_warnings": sum(
                    1
                    for item in items
                    if item.get(
                        "warnings"
                    )
                ),

                "with_errors": sum(
                    1
                    for item in items
                    if not item.get(
                        "success"
                    )
                ),

                "blocked": sum(
                    1
                    for item in items
                    if not item.get(
                        "ready_for_analysis"
                    )
                ),
            },

            "courses": items,
        }
        
    # ========================================================
    # PDF -> TEXTO
    # ========================================================

    @staticmethod
    def _pdf_to_text(
        content: bytes,
    ) -> str:

        reader = PdfReader(
            io.BytesIO(
                content
            )
        )

        pages: List[str] = []

        for index, page in enumerate(
            reader.pages,
            start=1,
        ):

            page_text = (
                page.extract_text()
                or ""
            ).strip()

            if page_text:

                pages.append(
                    f"--- Página {index} ---\n"
                    f"{page_text}"
                )

        return "\n\n".join(
            pages
        ).strip()

    # ========================================================
    # XLSX -> TEXTO
    # ========================================================

    @staticmethod
    def _xlsx_to_text(
        content: bytes,
    ) -> str:

        workbook = (
            openpyxl
            .load_workbook(
                io.BytesIO(
                    content
                ),
                read_only=True,
                data_only=True,
            )
        )

        parts: List[str] = []

        try:

            for worksheet in (
                workbook.worksheets
            ):

                parts.append(
                    "==============================\n"
                    f"HOJA: {worksheet.title}\n"
                    "=============================="
                )

                declared_rows = (
                    worksheet.max_row
                    or 1
                )

                declared_columns = (
                    worksheet.max_column
                    or 1
                )

                maximum_rows = min(
                    declared_rows,
                    MAX_XLSX_ROWS,
                )

                maximum_columns = min(
                    declared_columns,
                    MAX_XLSX_COLUMNS,
                )

                empty_rows = 0

                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=maximum_rows,
                    min_col=1,
                    max_col=maximum_columns,
                    values_only=True,
                ):

                    values = [
                        (
                            str(
                                value
                            ).strip()
                            if value
                            is not None
                            else ""
                        )
                        for value
                        in row
                    ]

                    while (
                        values
                        and not values[-1]
                    ):
                        values.pop()

                    if not any(
                        values
                    ):
                        empty_rows += 1

                        if (
                            empty_rows
                            >= STOP_AFTER_EMPTY_ROWS
                        ):
                            break

                        continue

                    empty_rows = 0

                    parts.append(
                        " | ".join(
                            values
                        )
                    )

        finally:
            workbook.close()

        return "\n".join(
            parts
        ).strip()

    # ========================================================
    # EXTRAER UN DOCUMENTO
    # ========================================================

    def _extract_source_text(
        self,
        file: Dict[str, Any],
    ) -> Tuple[
        str,
        str,
        List[str],
    ]:

        file_id = file[
            "id"
        ]

        file_name = file.get(
            "name",
            file_id,
        )

        mime_type = file.get(
            "mimeType",
            "",
        )

        warnings: List[str] = []

        # ----------------------------------------------------
        # Google Sheets
        # ----------------------------------------------------

        if (
            mime_type
            == GOOGLE_SHEET_MIME
        ):

            try:

                pdf_bytes = (
                    drive_service
                    .export_workspace_file(
                        file_id,
                        PDF_MIME,
                    )
                )

                pdf_text = (
                    self._pdf_to_text(
                        pdf_bytes
                    )
                )

                # Liberar lo antes posible.
                pdf_bytes = None

                if len(
                    pdf_text.strip()
                ) >= 50:

                    return (
                        pdf_text,
                        "pdf",
                        warnings,
                    )

                warnings.append(
                    f"{file_name}: el PDF temporal no "
                    "produjo suficiente texto. "
                    "Se utilizará XLSX como respaldo."
                )

            except Exception as exc:

                warnings.append(
                    f"{file_name}: no se pudo procesar "
                    f"el PDF temporal ({exc}). "
                    "Se utilizará XLSX como respaldo."
                )

            # Respaldo XLSX.
            xlsx_bytes = (
                drive_service
                .download_file(
                    file_id
                )
            )

            if not xlsx_bytes:

                raise CurriculumFeedbackError(
                    "No se pudo exportar "
                    f"'{file_name}' como XLSX"
                )

            xlsx_text = (
                self._xlsx_to_text(
                    xlsx_bytes
                )
            )

            xlsx_bytes = None

            return (
                xlsx_text,
                "xlsx_fallback",
                warnings,
            )

        # ----------------------------------------------------
        # PDF normal
        # ----------------------------------------------------

        if mime_type == PDF_MIME:

            content = (
                drive_service
                .download_file(
                    file_id
                )
            )

            if not content:

                raise CurriculumFeedbackError(
                    f"No se pudo descargar '{file_name}'"
                )

            return (
                self._pdf_to_text(
                    content
                ),
                "pdf",
                warnings,
            )

        # ----------------------------------------------------
        # XLSX normal
        # ----------------------------------------------------

        if mime_type in {
            XLSX_MIME,
            "application/vnd.ms-excel",
        }:

            content = (
                drive_service
                .download_file(
                    file_id
                )
            )

            if not content:

                raise CurriculumFeedbackError(
                    f"No se pudo descargar '{file_name}'"
                )

            return (
                self._xlsx_to_text(
                    content
                ),
                "xlsx",
                warnings,
            )

        raise CurriculumFeedbackError(
            f"Formato no soportado para "
            f"'{file_name}': {mime_type}"
        )

    # ========================================================
    # CARGAR CONTEXTO 1..5
    # ========================================================

    def load_context(
        self,
        files: Dict[
            str,
            Optional[
                Dict[str, Any]
            ],
        ],
    ) -> Tuple[
        Dict[str, Dict[str, Any]],
        List[str],
    ]:

        if not files.get(
            MAIN_INPUT_KEY
        ):

            raise CurriculumFeedbackError(
                "No se encontró 5_Diseño_Curricular. "
                "El análisis no puede continuar."
            )

        documents: Dict[
            str,
            Dict[str, Any],
        ] = {}

        warnings: List[str] = []

        for spec in PLANNING_FILES[:5]:

            file = files.get(
                spec.key
            )

            # ----------------------------------------------
            # Faltan documentos opcionales 1..4
            # ----------------------------------------------

            if not file:

                if (
                    spec.key
                    in OPTIONAL_CONTEXT_KEYS
                ):

                    warnings.append(
                        f"No se encontró {spec.name}. "
                        "El análisis continuará con "
                        "menos contexto."
                    )

                    continue

                raise CurriculumFeedbackError(
                    "No se encontró el archivo obligatorio "
                    f"{spec.name}"
                )

            # ----------------------------------------------
            # Extraer
            # ----------------------------------------------

            try:

                workbook_info = None

                if spec.key == MAIN_INPUT_KEY:

                    (
                        text,
                        extraction_method,
                        extraction_warnings,
                        workbook_info,
                    ) = self._extract_curriculum_source(
                        file
                    )

                else:

                    (
                        text,
                        extraction_method,
                        extraction_warnings,
                    ) = self._extract_source_text(
                        file
                    )

            except Exception as exc:

                if (
                    spec.key
                    in OPTIONAL_CONTEXT_KEYS
                ):

                    warnings.append(
                        f"No se pudo leer {spec.name}: "
                        f"{exc}. "
                        "El análisis continuará."
                    )

                    continue

                raise

            warnings.extend(
                extraction_warnings
            )

            if not text.strip():

                if (
                    spec.key
                    in OPTIONAL_CONTEXT_KEYS
                ):

                    warnings.append(
                        f"{spec.name} no contiene "
                        "texto utilizable. "
                        "El análisis continuará."
                    )

                    continue

                raise CurriculumFeedbackError(
                    "5_Diseño_Curricular no produjo "
                    "contenido utilizable"
                )

            documents[
                spec.key
            ] = {
                "name": (
                    file.get("name")
                    or spec.name
                ),

                "text": text,

                "chars": len(text),

                "extraction_method": (
                    extraction_method
                ),

                "workbook": (
                    workbook_info
                    if spec.key == MAIN_INPUT_KEY
                    else None
                ),
            }

        return (
            documents,
            warnings,
        )


    def _find_revision_folder(
        self,
        course_folder_id: str,
    ) -> Optional[Dict[str, Any]]:

        folders = (
            drive_service.list_folders(
                course_folder_id
            )
        )

        expected_names = {
            self.normalize(
                "0_Revision_de_Material"
            ),
            self.normalize(
                "Revision_de_Material"
            ),
        }

        matches = [
            folder
            for folder in folders
            if self.normalize(
                folder.get(
                    "name"
                )
            )
            in expected_names
        ]

        if len(matches) > 1:

            raise CurriculumFeedbackError(
                "Se encontraron varias carpetas "
                "0_Revision_de_Material"
            )

        return (
            matches[0]
            if matches
            else None
        )


    def _find_matrix_file(
        self,
        revision_folder_id: str,
    ) -> Optional[Dict[str, Any]]:

        files = (
            drive_service.list_files(
                revision_folder_id
            )
        )

        expected_names = {
            self.normalize(
                "02_Matriz observaciones estructura"
            ),
            self.normalize(
                "Matriz observaciones estructura"
            ),
        }

        matches = [
            file
            for file in files
            if self.normalize(
                file.get(
                    "name"
                )
            )
            in expected_names
        ]

        if len(matches) > 1:

            raise CurriculumFeedbackError(
                "Se encontraron varias matrices llamadas "
                "02_Matriz observaciones estructura"
            )

        if not matches:
            return None

        matrix = matches[0]

        if (
            matrix.get(
                "mimeType"
            )
            != GOOGLE_SHEET_MIME
        ):

            raise CurriculumFeedbackError(
                "02_Matriz observaciones estructura "
                "debe ser un Google Sheets nativo"
            )

        return matrix

    def _is_matrix_target(
        self,
        label: str,
    ) -> bool:

        normalized = self.normalize(
            label
        )

        fixed = {
            self.normalize(
                "Análisis Internacional y Local"
            ),
            self.normalize(
                "Competencias"
            ),
            self.normalize(
                "Diseño curricular"
            ),
            self.normalize(
                "Semana Diagnostico"
            ),
            self.normalize(
                "Semana de Diagnóstico"
            ),
        }

        if normalized in fixed:
            return True

        if re.fullmatch(
            r"semana\s+\d+",
            normalized,
        ):
            return True

        if re.fullmatch(
            r"proyecto\s+\d+",
            normalized,
        ):
            return True

        if re.fullmatch(
            r"practica\s+\d+",
            normalized,
        ):
            return True

        if normalized.startswith(
            "tarea "
        ):
            return True

        return False

    def _matrix_target_id(
        self,
        label: str,
    ) -> str:

        normalized = self.normalize(
            label
        )

        normalized = normalized.replace(
            ".",
            ""
        )

        target_id = re.sub(
            r"[^a-z0-9]+",
            "_",
            normalized,
        ).strip("_")

        return target_id

    def _matrix_targets(
        self,
        spreadsheet_id: str,
    ) -> Tuple[
        str,
        List[Dict[str, Any]],
    ]:

        sheet_title = (
            google_sheets_service
            .get_first_sheet_title(
                spreadsheet_id
            )
        )

        values = (
            google_sheets_service
            .get_column_values(
                spreadsheet_id,
                sheet_title,
                "A",
            )
        )

        targets: List[
            Dict[str, Any]
        ] = []

        target_ids = set()

        for row_number, row in enumerate(
            values,
            start=1,
        ):

            if not row:
                continue

            label = str(
                row[0]
                or ""
            ).strip()

            if not label:
                continue

            if not self._is_matrix_target(
                label
            ):
                continue

            target_id = (
                self._matrix_target_id(
                    label
                )
            )

            if target_id in target_ids:

                raise CurriculumFeedbackError(
                    "La matriz contiene una sección "
                    f"duplicada: {label}"
                )

            target_ids.add(
                target_id
            )

            targets.append(
                {
                    "id": target_id,
                    "label": label,
                    "row": row_number,
                }
            )

        if not targets:

            raise CurriculumFeedbackError(
                "No se encontraron secciones "
                "utilizables en la matriz"
            )

        return (
            sheet_title,
            targets,
        )

    # ========================================================
    # ARMAR CONTEXTO
    # ========================================================

    def _context_text(
        self,
        documents: Dict[
            str,
            Dict[str, Any],
        ],
    ) -> Tuple[str, bool]:
        """
        El documento 5 SIEMPRE tiene prioridad.

        Si por seguridad debemos recortar contexto:
        1. conservamos primero 5_Diseño_Curricular;
        2. repartimos el espacio restante entre 1..4.
        """
        main_document = documents[
            MAIN_INPUT_KEY
        ]

        main_block = (
            "========================================\n"
            "DOCUMENTO PRINCIPAL\n"
            f"{main_document['name']}\n"
            "========================================\n"
            f"{main_document['text']}"
        )

        if (
            len(main_block)
            >= MAX_CONTEXT_CHARS
        ):

            return (
                main_block[
                    :MAX_CONTEXT_CHARS
                ],
                True,
            )

        optional_blocks: List[
            str
        ] = []

        for spec in PLANNING_FILES[:4]:

            document = documents.get(
                spec.key
            )

            if not document:
                continue

            block = (
                "========================================\n"
                "DOCUMENTO DE CONTEXTO\n"
                f"{document['name']}\n"
                "========================================\n"
                f"{document['text']}"
            )

            optional_blocks.append(
                block
            )

        if not optional_blocks:

            return (
                main_block,
                False,
            )

        remaining = (
            MAX_CONTEXT_CHARS
            - len(main_block)
            - 10
        )

        complete_length = sum(
            len(block)
            for block
            in optional_blocks
        )

        if (
            complete_length
            <= remaining
        ):

            return (
                main_block
                + "\n\n"
                + "\n\n".join(
                    optional_blocks
                ),
                False,
            )

        # Repartir equitativamente el espacio disponible
        # para no eliminar completamente un documento de contexto.
        per_document = max(
            1,
            remaining
            // len(optional_blocks)
        )

        trimmed_blocks = [
            block[
                :per_document
            ]
            for block
            in optional_blocks
        ]

        return (
            main_block
            + "\n\n"
            + "\n\n".join(
                trimmed_blocks
            ),
            True,
        )

    def _matrix_response_schema(
        self,
        targets: List[
            Dict[str, Any]
        ],
    ) -> Dict[str, Any]:

        properties = {
            target["id"]: {
                "type": "STRING",
            }
            for target
            in targets
        }

        required = [
            target["id"]
            for target
            in targets
        ]

        return {
            "type": "OBJECT",

            "properties": {
                "observaciones": {
                    "type": "OBJECT",
                    "properties": properties,
                    "required": required,
                },

                "resumen_general": {
                    "type": "STRING",
                },

                "advertencias": {
                    "type": "ARRAY",
                    "items": {
                        "type": "STRING",
                    },
                },
            },

            "required": [
                "observaciones",
                "resumen_general",
                "advertencias",
            ],
        }

    # ========================================================
    # PROMPT
    # ========================================================

    def _build_prompt(
        self,
        course: CourseCatalog,
        context_text: str,
        warnings: List[str],
        matrix_targets: List[
            Dict[str, Any]
        ],
    ) -> Tuple[
        str,
        str,
    ]:

        # ========================================================
        # SYSTEM MESSAGE
        # ========================================================

        system_message = """
    Eres un especialista en revisión crítica de diseño curricular
    universitario.

    Tu tarea es analizar la COHERENCIA, PERTINENCIA y SENTIDO
    ACADÉMICO de una Planeación Curricular.

    NO estás realizando:

    - un resumen;
    - un FODA;
    - una validación mecánica de celdas;
    - una revisión de formato de Excel;
    - una revisión de filas o coordenadas;
    - una descripción superficial del contenido.

    Debes actuar como revisor curricular crítico.

    Tu pregunta principal durante todo el análisis debe ser:

    "¿Lo que fue escrito aquí tiene sentido académico y mantiene
    coherencia con lo que el curso pretende enseñar, con las
    competencias, con los contenidos, con las actividades, con
    los tiempos y con el contexto proporcionado?"

    ============================================================
    PRINCIPIO FUNDAMENTAL
    ============================================================

    ANTES de evaluar cualquier apartado:

    1. Determina si realmente existe contenido proporcionado
    por el autor.

    2. Identifica qué pretende hacer ese apartado.

    3. Identifica con qué competencia, contenido, actividad,
    recurso, proyecto, práctica o tarea se relaciona.

    4. Contrástalo con las fuentes pertinentes.

    5. Analiza si existe coherencia y sentido.

    6. Solo después genera la observación.

    NO debes concluir que algo es correcto simplemente porque:

    - tiene información;
    - una fórmula indica "VALIDACIÓN CORRECTA";
    - el tiempo suma correctamente;
    - existe una competencia escrita;
    - existe una descripción;
    - existe una URL;
    - existe un título.

    Todo debe tener SENTIDO en conjunto.

    ============================================================
    FUENTES
    ============================================================

    La fuente principal de la Planeación Curricular es:

    5_Diseño_Curricular

    Sus hojas contienen:

    - Competencias;
    - Diseño;
    - Semana Diagnostico;
    - S.2 a S.11;
    - Proyectos;
    - Practicas;
    - Tareas.

    También existen documentos de contexto:

    1_Fortalezas_Debilidades_y_Recomendaciones
    2_Analisis_de_Contexto
    3_Criterios y Expectativas
    4_Analisis_Internacional

    Cada documento tiene una función distinta.

    NO mezcles problemas encontrados en una fuente con otra sección.

    ============================================================
    AISLAMIENTO DE HALLAZGOS
    ============================================================

    Un problema pertenece ÚNICAMENTE al apartado donde realmente
    fue encontrado.

    Ejemplos:

    - Un problema encontrado en S.5 pertenece a semana_5.
    - Un problema encontrado en Proyecto 1 pertenece a proyecto_1.
    - Un problema encontrado en Practica 1 pertenece a practica_1.
    - Un problema encontrado en una tarea pertenece a esa tarea.

    NUNCA traslades un error de una semana hacia:

    analisis_internacional_y_local

    NUNCA traslades un error de un proyecto hacia una semana.

    NUNCA traslades un error de una práctica hacia Competencias.

    Cada observación debe estar sustentada por evidencia del
    apartado que se está evaluando o por una comparación válida
    entre ese apartado y su fuente de referencia.

    ============================================================
    COORDENADAS, FILAS Y CELDAS
    ============================================================

    El contenido recibido puede contener referencias como:

    A1
    B28
    C15
    fila 20
    celda B28
    hoja S.5

    Estas referencias existen únicamente porque el backend
    preserva la estructura del documento.

    NO debes utilizar coordenadas, números de fila ni números de
    celda en las observaciones finales.

    INCORRECTO:

    "Se detectó un error #REF! en la celda B28."

    CORRECTO:

    "2.1 Contenido - Unidad: se detecta una referencia inválida
    en el valor de la unidad, por lo que no es posible comprobar
    su correspondencia con el tema programado."

    La observación debe identificar el ELEMENTO CURRICULAR,
    no su ubicación física en Excel.

    ============================================================
    TIPO DE ERRORES QUE INTERESAN
    ============================================================

    Los hallazgos principales deben ser errores o inconsistencias
    de COHERENCIA y SENTIDO.

    Debes buscar principalmente:

    - competencia que no corresponde con el contenido;
    - contenido que no corresponde con la semana;
    - tema que no corresponde con la unidad;
    - subtema que no corresponde con el tema;
    - valor actitudinal sin relación real con la semana;
    - descripción genérica que no explica cómo se aplicará;
    - actividad incompatible con la competencia;
    - tipo de actividad incompatible con su descripción;
    - tiempo poco razonable para lo que se pretende realizar;
    - recurso educativo que no tiene sentido para ese contenido;
    - tiempo del recurso que no guarda relación con su uso;
    - proyecto sin relación suficiente con el curso;
    - proyecto cuya competencia no corresponde con lo solicitado;
    - tecnología que no corresponde con la descripción;
    - práctica fuera de contexto;
    - tarea que no refuerza conocimientos necesarios;
    - secuencia de contenidos poco lógica;
    - actividades que requieren conocimientos todavía no desarrollados;
    - cierres que asignan actividades inexistentes o fuera de secuencia;
    - ponderaciones poco coherentes con el trabajo solicitado;
    - descripciones que contradicen el título, competencia,
    herramientas o propósito de la actividad.

    Los límites de tiempo establecidos en la plantilla también
    deben comprobarse, pero NO conviertas el análisis en una
    simple validación numérica.

    El tiempo también debe tener SENTIDO respecto de lo que se
    pretende realizar.

    ============================================================
    ESTILO
    ============================================================

    NO uses:

    "Fortaleza:"
    "Debilidad:"
    "Aspecto positivo:"
    "Aspecto negativo:"

    NO escribas observaciones vacías como:

    "Todo está correcto."

    "Existe coherencia."

    "El contenido está bien estructurado."

    "La planificación es adecuada."

    "Cumple con lo esperado."

    Si no encuentras problemas, debes mencionar concretamente
    qué relaciones revisaste.

    Ejemplo aceptable:

    "No se identifican inconsistencias entre la competencia
    seleccionada, los contenidos de la semana, las actividades
    propuestas y los tiempos asignados."

    ============================================================
    ERRORES TÉCNICOS
    ============================================================

    Errores como:

    #REF!
    #VALUE!
    #N/A
    #DIV/0!
    #NAME?

    solo deben mencionarse cuando aparezcan REALMENTE dentro del
    elemento curricular que estás analizando.

    Nunca inventes un error técnico.

    Nunca asumas que un error encontrado en una parte del documento
    está presente en otra.

    ============================================================
    SALIDA
    ============================================================

    Las observaciones se escribirán directamente en:

    "Observaciones Generales por la IA"

    Por ello:

    - deben ser claras;
    - deben ser específicas;
    - deben indicar el elemento revisado;
    - deben explicar el motivo;
    - no deben depender de coordenadas del Excel.

    Responde únicamente JSON válido.
    """.strip()

        # ========================================================
        # BACKEND WARNINGS
        # ========================================================

        warning_text = (
            "\n".join(
                f"- {warning}"
                for warning in warnings
            )
            or "- Ninguna"
        )

        # ========================================================
        # DYNAMIC MATRIX TARGETS
        # ========================================================

        target_text = "\n".join(
            (
                f'- {target["id"]}: '
                f'{target["label"]}'
            )
            for target
            in matrix_targets
        )

        example_fields = ",\n".join(
            (
                f'    "{target["id"]}": '
                '"observación correspondiente"'
            )
            for target
            in matrix_targets
        )

        # ========================================================
        # USER MESSAGE
        # ========================================================

        user_message = f"""
    ============================================================
    CURSO
    ============================================================

    Código: {course.code}
    Nombre: {course.name}


    ============================================================
    INFORMACIÓN OPERATIVA DEL BACKEND
    ============================================================

    {warning_text}

    IMPORTANTE:

    Estas advertencias son información OPERATIVA del backend.

    NO constituyen evidencia curricular.

    NO copies estas advertencias dentro de las observaciones
    curriculares salvo que indiquen explícitamente que una fuente
    necesaria no pudo ser leída.

    No conviertas mensajes técnicos del backend en errores del
    contenido curricular.


    ============================================================
    OBJETIVO GENERAL
    ============================================================

    Realiza una revisión CRÍTICA de la Planeación Curricular.

    El objetivo principal es determinar si cada elemento:

    - existe cuando corresponde;
    - tiene sentido;
    - mantiene coherencia interna;
    - mantiene coherencia con su competencia;
    - mantiene coherencia con los contenidos;
    - mantiene coherencia con las actividades;
    - mantiene coherencia con los tiempos;
    - mantiene coherencia con los documentos de contexto;
    - mantiene coherencia con la progresión general del curso.

    NO busques errores únicamente por encontrar errores.

    Pero tampoco aceptes información superficialmente solo porque
    está completa.

    Debes cuestionar críticamente si lo escrito tiene SENTIDO.


    ============================================================
    MÉTODO OBLIGATORIO PARA TODO APARTADO
    ============================================================

    Para cada sección sigue este orden:

    PASO 1 — PRESENCIA

    Determina si existe contenido REAL.

    No consideres contenido real únicamente:

    - títulos;
    - etiquetas;
    - campos vacíos;
    - "Obligatorio";
    - "Obligatoria";
    - "Optativa";
    - nombres de plantilla.

    PASO 2 — PROPÓSITO

    Identifica qué pretende lograr el apartado.

    PASO 3 — REFERENCIA

    Determina contra qué debe compararse:

    - Competencias;
    - Diseño;
    - contexto;
    - semana;
    - proyecto;
    - práctica;
    - tarea;
    - reglas de tiempo;
    - otros elementos relacionados.

    PASO 4 — COHERENCIA

    Analiza si las piezas tienen sentido entre sí.

    PASO 5 — OBSERVACIÓN

    Escribe únicamente hallazgos específicos y útiles.


    ============================================================
    ANÁLISIS INTERNACIONAL Y LOCAL
    ============================================================

    CLAVE:

    analisis_internacional_y_local

    FUENTE PRINCIPAL Y OBLIGATORIA PARA ESTA OBSERVACIÓN:

    4_Analisis_Internacional

    Si dentro del contexto aparece identificado como:

    "Analisis Internacional"
    "Análisis Internacional"
    "Analisis Internacional y Local"
    "Análisis Internacional y Local"

    trátalo como la misma fuente.

    ANTES de generar esta observación debes LEER específicamente
    ese documento.

    NO utilices como evidencia principal:

    - S.2;
    - S.3;
    - S.4;
    - S.5;
    - S.6;
    - S.7;
    - S.8;
    - S.9;
    - S.10;
    - S.11;
    - Proyectos;
    - Practicas;
    - Tareas.

    Estas secciones pueden utilizarse únicamente después para
    comprobar si lo planteado en el análisis internacional/local
    tiene correspondencia general con el Diseño Curricular.

    MUY IMPORTANTE:

    Un #REF!, tiempo incorrecto, actividad incorrecta o cualquier
    otro error encontrado en una semana NO pertenece a
    analisis_internacional_y_local.

    No menciones ese error aquí.

    PARA ESTA SECCIÓN ANALIZA:

    1. Qué referentes internacionales o locales se utilizaron.

    2. Qué tendencias, contenidos, competencias, metodologías,
    tecnologías o enfoques fueron identificados.

    3. Si el análisis establece una comparación real y útil.

    4. Si las conclusiones obtenidas tienen sentido respecto del
    curso.

    5. Si existe coherencia entre lo identificado internacional
    y localmente y las decisiones curriculares generales.

    6. Si las recomendaciones o conclusiones del análisis fueron
    incorporadas de manera razonable en el Diseño Curricular.

    7. Si existen conclusiones que no se derivan de la evidencia
    presentada.

    8. Si existen comparaciones superficiales que no sustentan una
    decisión curricular.

    La observación debe tratar sobre la CALIDAD Y COHERENCIA DEL
    ANÁLISIS INTERNACIONAL Y LOCAL.

    Ejemplo:

    "Se identifican referentes internacionales relacionados con
    almacenamiento distribuido y virtualización, y estos guardan
    correspondencia con los contenidos incorporados al Diseño
    Curricular. No se identifican discrepancias relevantes entre
    los referentes analizados y la orientación general del curso."

    O, si existe un problema real:

    "Las tendencias identificadas priorizan X, pero esta conclusión
    no se refleja posteriormente en las competencias ni en el
    Diseño Curricular, por lo que existe una desconexión entre el
    análisis realizado y la planificación final."

    NO menciones filas.
    NO menciones celdas.
    NO menciones errores de otras hojas.


    ============================================================
    COMPETENCIAS
    ============================================================

    CLAVE:

    competencias

    FUENTE PRINCIPAL:

    hoja Competencias de 5_Diseño_Curricular.

    Primero identifica cuáles son las competencias OFICIALES del
    curso.

    Estas competencias serán la referencia principal para:

    - semanas;
    - proyectos;
    - prácticas;
    - tareas.

    Analiza críticamente cada competencia considerando:

    - verbo utilizado;
    - objeto de aprendizaje;
    - conocimientos implicados;
    - habilidad esperada;
    - condición o contexto de aplicación;
    - herramientas cuando corresponda;
    - relación con el perfil y contexto del curso.

    Busca especialmente:

    - competencias demasiado genéricas;
    - competencias que no corresponden con el curso;
    - competencias redundantes;
    - competencias que mezclan demasiados resultados distintos;
    - competencias que no pueden evidenciarse mediante las
    actividades propuestas;
    - competencias que no aparecen luego en el Diseño Curricular;
    - contenidos importantes que no tienen una competencia
    relacionada.

    No evalúes únicamente la redacción.

    Evalúa el SENTIDO curricular de las competencias.


    ============================================================
    DISEÑO CURRICULAR
    ============================================================

    CLAVE:

    diseno_curricular

    FUENTES PRINCIPALES:

    - hoja Diseño;
    - hoja Competencias.

    Utiliza además los documentos contextuales para comprobar
    pertinencia.

    Analiza globalmente:

    - unidades;
    - temas;
    - subtemas;
    - competencias;
    - secuencia;
    - distribución semanal;
    - progresión;
    - relación entre teoría y práctica;
    - relación con proyectos;
    - relación con prácticas;
    - relación con tareas.

    Busca:

    - temas fuera de secuencia;
    - subtemas que no corresponden con su tema;
    - unidades desconectadas;
    - contenidos repetidos sin propósito;
    - saltos de dificultad;
    - contenidos que requieren conocimientos no desarrollados;
    - contenido que no responde a las competencias;
    - competencias sin contenido suficiente;
    - contenidos sin competencia relacionada.

    Aquí debes señalar problemas GLOBALES.

    Los errores específicos de una semana deben permanecer en esa
    semana.


    ============================================================
    SEMANA DIAGNOSTICO
    ============================================================

    CLAVE:

    semana_diagnostico

    La Semana Diagnóstico debe analizarse como una herramienta para
    determinar el punto de partida de los estudiantes.

    PRIMERO comprueba qué apartados realmente contienen
    información.

    DESPUÉS analiza:

    ------------------------------------------------------------
    1. Bienvenida y Presentación
    ------------------------------------------------------------

    Revisa:

    - qué se pretende presentar;
    - cantidad de información;
    - propósito;
    - tiempo.

    Tiempo máximo indicado por la plantilla:

    15 minutos.

    No basta con comprobar el límite.

    Determina también si lo descrito puede desarrollarse
    razonablemente en ese tiempo.


    ------------------------------------------------------------
    2. Evaluación de Conocimientos Previos
    ------------------------------------------------------------

    Revisa:

    - conocimientos que se pretenden diagnosticar;
    - relación con los conocimientos necesarios para el curso;
    - preguntas;
    - parte práctica;
    - tiempo.

    La plantilla contempla aproximadamente:

    15 a 20 preguntas
    +
    una parte práctica.

    Tiempo máximo:

    45 minutos.

    Sé crítico:

    los conocimientos evaluados deben ser conocimientos que ayuden
    realmente a determinar si el estudiante posee bases para el
    curso.

    No deben evaluar arbitrariamente contenido futuro sin sentido
    diagnóstico.


    ------------------------------------------------------------
    3. Presentación de los Estudiantes
    ------------------------------------------------------------

    Comprueba:

    - propósito;
    - metodología;
    - preguntas;
    - cantidad de estudiantes;
    - tiempo.

    La plantilla plantea aproximadamente 5 estudiantes.

    Tiempo máximo:

    30 minutos.

    Analiza también si la dinámica descrita tiene sentido dentro
    de ese tiempo.


    ------------------------------------------------------------
    4. Censo de Horarios de Calificación
    ------------------------------------------------------------

    Comprueba:

    - propósito;
    - mecanismo;
    - horarios;
    - formulario cuando corresponda;
    - tiempo.

    Referencia aproximada:

    5 minutos.


    ------------------------------------------------------------
    5. Presentación del Programa del Curso
    ------------------------------------------------------------

    Comprueba:

    - propósito;
    - información que se pretende explicar;
    - tiempo;
    - coherencia con el inicio del curso.

    Tiempo máximo:

    30 minutos.


    ------------------------------------------------------------
    6. Evaluación de Conocimientos del Curso Actual
    ------------------------------------------------------------

    Comprueba:

    - qué pretende evaluar;
    - preguntas;
    - relación con el propósito diagnóstico;
    - tiempo.

    La plantilla plantea aproximadamente:

    20 preguntas.

    Tiempo máximo:

    15 minutos.


    ------------------------------------------------------------
    COHERENCIA GLOBAL DE SEMANA DIAGNOSTICO
    ------------------------------------------------------------

    Después revisa:

    - si los bloques realmente permiten diagnosticar al estudiante;
    - si existe redundancia;
    - si falta evaluar una base importante;
    - si los tiempos tienen sentido;
    - si el conjunto de actividades cabe razonablemente en el
    tiempo disponible;
    - si los conocimientos evaluados son pertinentes para el curso.

    FORMATO CUANDO HAYA ERRORES:

    1. Evaluación de Conocimientos Previos
    Contenido: ...
    Tiempo: ...

    3. Presentación de los Estudiantes
    Metodología: ...

    Incluye únicamente los apartados que requieren observación.

    No menciones filas ni celdas.


    ============================================================
    SEMANAS S.2 A S.11
    ============================================================

    CLAVES:

    semana_2
    semana_3
    semana_4
    semana_5
    semana_6
    semana_7
    semana_8
    semana_9
    semana_10
    semana_11

    Cada semana debe analizarse DE FORMA INDEPENDIENTE.

    No reutilices automáticamente una observación entre semanas.

    Antes de evaluar la semana:

    1. identifica en la hoja Diseño qué contenido corresponde a
    esa semana;

    2. identifica qué competencia corresponde;

    3. identifica cuál es la progresión proveniente de semanas
    anteriores;

    4. después revisa el contenido detallado de la hoja S.X.


    ============================================================
    1. ÁREA ACTITUDINAL — SABER SER
    ============================================================

    Para cada semana analiza:

    - Nombre del Valor;
    - Tiempo Aproximado;
    - Descripción;
    - relación del valor con los conocimientos de esa semana;
    - explicación de cómo se aplicará.


    ------------------------------------------------------------
    VALOR
    ------------------------------------------------------------

    No basta con que exista un valor como:

    Responsabilidad
    Trabajo en equipo
    Integridad
    Proactividad
    Adaptabilidad
    Colaboración

    Debes determinar si ESE valor tiene sentido respecto de:

    - contenidos de esa semana;
    - actividades;
    - dinámica de trabajo;
    - competencia;
    - contexto.

    Ejemplo de problema:

    "1. Área Actitudinal (Saber Ser)
    Valor: se selecciona Trabajo en Equipo, pero las actividades
    de la semana son completamente individuales y la descripción
    no explica ninguna dinámica colaborativa."


    ------------------------------------------------------------
    DESCRIPCIÓN
    ------------------------------------------------------------

    Debe explicar cómo se desarrolla el valor.

    Una definición genérica de "responsabilidad" o "integridad"
    no es suficiente.

    Debe existir relación con lo que el estudiante realizará.


    ------------------------------------------------------------
    RELACIÓN CON LOS CONOCIMIENTOS
    ------------------------------------------------------------

    Comprueba que la explicación conecte de forma concreta el
    valor con los conocimientos de ESA semana.

    Si menciona contenidos de otra semana o utiliza una explicación
    genérica sin conexión, indícalo.


    ------------------------------------------------------------
    TIEMPO
    ------------------------------------------------------------

    La plantilla establece aproximadamente:

    mínimo 3 minutos
    máximo 5 minutos.

    Comprueba:

    - que se encuentre en el rango;
    - que además sea suficiente para lo descrito.

    No evalúes únicamente el número.


    ============================================================
    2. ÁREA CONOCIMIENTO — SABER
    ============================================================

    Analiza:

    - Tiempo Aproximado;
    - Competencia a desarrollar;
    - Contenido;
    - Recursos Educativos.


    ------------------------------------------------------------
    COMPETENCIA
    ------------------------------------------------------------

    La competencia utilizada debe compararse contra la hoja
    Competencias.

    Debe:

    1. existir en Competencias;
    2. conservar el mismo sentido;
    3. tener relación con el contenido de esa semana.

    Ignora únicamente diferencias de:

    - espacios;
    - saltos de línea;
    - mayúsculas/minúsculas;
    - formato.

    No aceptes una competencia diferente simplemente porque
    "se parece".

    Ejemplo:

    "2. Área Conocimiento (Saber)
    Competencia: la competencia utilizada no corresponde con las
    competencias definidas para el curso y tampoco guarda relación
    suficiente con el contenido programado para esta semana."


    ------------------------------------------------------------
    TIEMPO
    ------------------------------------------------------------

    El Área Conocimiento no debe superar aproximadamente:

    60 minutos.

    Pero debes analizar también si el tiempo asignado tiene
    sentido para:

    - cantidad de contenidos;
    - profundidad;
    - recursos;
    - complejidad.


    ============================================================
    2.1 CONTENIDO
    ============================================================

    Analiza por separado:

    - Unidad;
    - Tema;
    - Subtema.

    PRIMERA COMPARACIÓN:

    hoja Diseño.

    La Unidad, Tema y Subtema de la semana deben corresponder con
    lo programado para ESA semana.

    SEGUNDA COMPARACIÓN:

    competencia.

    El contenido debe contribuir realmente al desarrollo de la
    competencia.

    TERCERA COMPARACIÓN:

    contexto y secuencia.

    El contenido debe tener sentido respecto de:

    - semanas anteriores;
    - semanas posteriores;
    - conocimientos previos;
    - contexto del curso.

    Busca críticamente:

    - unidad de otra semana;
    - tema de otra unidad;
    - subtema sin relación con el tema;
    - contenido demasiado avanzado para ese momento;
    - contenido ya desarrollado innecesariamente;
    - contenido ajeno al curso;
    - contenido que no aporta a la competencia;
    - falta de un conocimiento necesario antes de avanzar.

    FORMATO:

    2.1 Contenido
    Unidad: ...
    Tema: ...
    Subtema: ...

    Solo incluye los campos que presenten una observación.


    ============================================================
    2.2 RECURSOS EDUCATIVOS
    ============================================================

    Analiza cada recurso REALMENTE utilizado.

    La plantilla contempla:

    Presentación
    - obligatorio;
    - máximo recomendado aproximado: 30 minutos.

    Video
    - obligatorio;
    - máximo recomendado aproximado: 6 minutos.

    Lectura
    - optativa;
    - máximo recomendado aproximado: 10 minutos.

    Ejercicio Demostrativo
    - obligatorio;
    - máximo recomendado aproximado: 30 minutos.

    Para cada recurso analiza:

    1. si tiene relación con el contenido;

    2. si es un recurso apropiado para lo que se pretende enseñar;

    3. si el tiempo tiene sentido;

    4. si el tiempo respeta la referencia de la plantilla;

    5. si la combinación de recursos puede realizarse dentro del
    tiempo total del Área Conocimiento;

    6. si existe redundancia entre recursos;

    7. si el recurso realmente aporta algo distinto.

    Ejemplo:

    "2.2 Recursos Educativos
    Video - Tiempo: se asignan 9 minutos cuando la plantilla
    recomienda un máximo de 6 minutos."

    Ejemplo de coherencia:

    "2.2 Recursos Educativos
    Lectura: el recurso se orienta a un contenido distinto al
    tema desarrollado durante la semana y no aporta directamente
    a la competencia seleccionada."


    ============================================================
    3. ÁREA DE HABILIDADES — SABER HACER
    ============================================================

    Analiza:

    - tiempo total;
    - Actividad 1;
    - Actividad 2 cuando exista.


    ============================================================
    ACTIVIDAD 1
    ============================================================

    La Actividad 1 debe contener:

    - Competencia que predomina;
    - Tipo de Actividad;
    - Ponderación;
    - Tiempo Aproximado;
    - Descripción de la Actividad.


    ------------------------------------------------------------
    COMPETENCIA
    ------------------------------------------------------------

    Comprueba:

    - que exista en Competencias;
    - que corresponda con el conocimiento trabajado;
    - que pueda desarrollarse mediante la actividad descrita.


    ------------------------------------------------------------
    TIPO DE ACTIVIDAD
    ------------------------------------------------------------

    El tipo de actividad debe tener SENTIDO con la descripción.

    Analiza semánticamente lo que el estudiante realmente hará.

    Ejemplo:

    Si se selecciona:

    "Actividad práctica"

    pero la descripción únicamente pide:

    "leer un documento y responder preguntas"

    existe una inconsistencia entre tipo y descripción.


    ------------------------------------------------------------
    DESCRIPCIÓN DE LA ACTIVIDAD
    ------------------------------------------------------------

    Este campo debe analizarse críticamente.

    Comprueba:

    - qué hará exactamente el estudiante;
    - qué conocimiento aplica;
    - qué competencia desarrolla;
    - si corresponde con el tipo de actividad;
    - si tiene relación con la semana;
    - si puede realizarse en el tiempo dado;
    - si utiliza conocimientos que ya deberían haberse aprendido;
    - si no exige conocimientos que todavía no se han desarrollado;
    - si el producto esperado tiene sentido.


    ------------------------------------------------------------
    TIEMPO
    ------------------------------------------------------------

    No te limites a comparar un rango.

    Analiza si el tiempo es REALISTA para la actividad descrita.

    Una actividad compleja con un tiempo mínimo puede ser
    incoherente aunque no exista una regla numérica explícita.


    ------------------------------------------------------------
    PONDERACIÓN
    ------------------------------------------------------------

    Analiza si la ponderación guarda una proporción razonable con:

    - dificultad;
    - tiempo;
    - producto;
    - importancia académica.

    No inventes una ponderación ideal si no existe evidencia
    suficiente.


    ============================================================
    ACTIVIDAD 2
    ============================================================

    PRIMERO determina si realmente existe.

    Si únicamente aparecen etiquetas de plantilla:

    no existe.

    Si está vacía:

    NO la menciones.

    Si existe, aplica exactamente la misma revisión de Actividad 1:

    - competencia;
    - tipo;
    - descripción;
    - tiempo;
    - ponderación;
    - coherencia con la semana.


    ============================================================
    CIERRE
    ============================================================

    Analiza:

    - Tiempo Aproximado;
    - Asignación de Actividades.

    Referencia aproximada:

    10 a 15 minutos.

    Comprueba si:

    - el tiempo tiene sentido;
    - permite cerrar correctamente la semana;
    - se relaciona con lo trabajado;
    - las actividades asignadas existen;
    - las tareas/proyectos/prácticas mencionados existen;
    - la asignación ocurre en una secuencia lógica;
    - no se asigna una actividad que exige contenido futuro.


    ============================================================
    COHERENCIA GLOBAL DE CADA SEMANA
    ============================================================

    Finalmente relaciona:

    VALOR
    ↓
    COMPETENCIA
    ↓
    CONTENIDO
    ↓
    RECURSOS
    ↓
    ACTIVIDAD
    ↓
    CIERRE

    La semana debe formar una unidad coherente.

    Pregúntate:

    "¿Tiene sentido que este valor, esta competencia, este
    contenido, estos recursos y esta actividad estén juntos en
    esta semana?"

    También analiza el tiempo total.

    La plantilla espera aproximadamente al menos 100 minutos
    totales y establece diferentes referencias por área.

    No confíes exclusivamente en una fórmula de validación.


    ============================================================
    FORMATO DE OBSERVACIONES DE SEMANAS
    ============================================================

    Si hay varios hallazgos:

    1. Área Actitudinal (Saber Ser)
    Valor: ...
    Tiempo: ...
    Relación con los conocimientos: ...

    2. Área Conocimiento (Saber)
    Competencia: ...

    2.1 Contenido
    Unidad: ...
    Tema: ...

    2.2 Recursos Educativos
    Video: ...

    3. Área de Habilidades (Saber Hacer)
    Actividad 1 - Tipo de Actividad: ...
    Actividad 1 - Descripción: ...
    Actividad 2 - Tiempo: ...

    Cierre
    Asignación de Actividades: ...

    NO escribas una sección que no tenga ninguna observación.

    NO menciones celdas.

    NO menciones filas.

    NO menciones coordenadas.


    ============================================================
    PROYECTOS
    ============================================================

    Correspondencia:

    Primer Proyecto -> proyecto_1
    Segundo Proyecto -> proyecto_2
    Tercer Proyecto -> proyecto_3

    ANTES DE ANALIZAR:

    determina si el proyecto realmente existe.

    No consideres contenido suficiente únicamente:

    - Primer Proyecto;
    - Segundo Proyecto;
    - Tercer Proyecto;
    - Obligatorio;
    - Optativa;
    - Competencia a desarrollar;
    - Título;
    - Tecnologías;
    - Descripción;
    - Ponderación;
    - Horas.

    Estas pueden ser únicamente etiquetas de plantilla.

    Si un proyecto optativo está vacío:

    devuelve exactamente "".


    ============================================================
    ANÁLISIS DE CADA PROYECTO EXISTENTE
    ============================================================

    Analiza como un CONJUNTO:

    COMPETENCIA
    ↓
    TÍTULO
    ↓
    TECNOLOGÍAS
    ↓
    DESCRIPCIÓN
    ↓
    PONDERACIÓN
    ↓
    HORAS

    No los revises aisladamente.


    ------------------------------------------------------------
    COMPETENCIA
    ------------------------------------------------------------

    Debe:

    - existir en Competencias;
    - corresponder realmente con el proyecto;
    - poder demostrarse mediante el producto solicitado.


    ------------------------------------------------------------
    TÍTULO
    ------------------------------------------------------------

    Debe representar adecuadamente lo que realmente se realizará.


    ------------------------------------------------------------
    TECNOLOGÍAS
    ------------------------------------------------------------

    Comprueba:

    - si tienen relación con el proyecto;
    - si son suficientes;
    - si la descripción realmente las utiliza;
    - si tienen sentido dentro del contexto del curso.


    ------------------------------------------------------------
    DESCRIPCIÓN
    ------------------------------------------------------------

    Sé especialmente crítico.

    Comprueba:

    - coherencia con competencia;
    - coherencia con título;
    - coherencia con tecnologías;
    - coherencia con los contenidos del curso;
    - coherencia con el contexto dado;
    - alcance;
    - dificultad;
    - secuencia;
    - producto esperado.

    Busca requisitos que parezcan desconectados del propósito
    principal.


    ------------------------------------------------------------
    HORAS
    ------------------------------------------------------------

    Comprueba si las horas tienen sentido respecto de:

    - alcance;
    - complejidad;
    - tecnologías;
    - producto.


    ------------------------------------------------------------
    PONDERACIÓN
    ------------------------------------------------------------

    Comprueba si tiene sentido respecto del trabajo solicitado.

    Además, cuando haya información suficiente, revisa la
    distribución global de proyectos según las reglas contenidas
    en la propia plantilla.


    ============================================================
    FORMATO DE PROYECTO
    ============================================================

    Ejemplo:

    Competencia: ...

    Descripción: ...

    Tecnologías: ...

    Horas: ...

    Incluye únicamente los aspectos donde exista una observación
    real.


    ============================================================
    PRACTICAS
    ============================================================

    Correspondencia:

    Primera Practica -> practica_1
    Segunda Practica -> practica_2
    Tercera Practica -> practica_3
    Cuarta Practica -> practica_4
    Quinta Practica -> practica_5
    Sexta Practica -> practica_6
    Septima Practica -> practica_7

    PRIMERO determina si cada práctica realmente tiene contenido.

    Si una práctica optativa está vacía:

    devuelve "".

    Para cada práctica existente analiza:

    COMPETENCIA
    ↓
    TÍTULO
    ↓
    TECNOLOGÍAS
    ↓
    DESCRIPCIÓN
    ↓
    PONDERACIÓN
    ↓
    HORAS

    La revisión debe enfocarse en SENTIDO y COHERENCIA.

    Pregúntate:

    - ¿la práctica desarrolla realmente la competencia indicada?;
    - ¿el título corresponde con lo que se hará?;
    - ¿las tecnologías corresponden?;
    - ¿la descripción aplica conocimientos del curso?;
    - ¿el estudiante posee esos conocimientos en ese momento?;
    - ¿el alcance tiene sentido?;
    - ¿las horas son razonables?;
    - ¿la ponderación guarda relación con el trabajo?;
    - ¿la práctica aporta algo distinto de un proyecto o tarea?;
    - ¿tiene sentido dentro del contexto del curso?

    Cuando exista información suficiente, comprueba también las
    reglas globales de ponderación indicadas en la plantilla.


    ============================================================
    TAREAS
    ============================================================

    Correspondencia:

    Tarea de fortalecimiento académico
    -> tarea_fortalecimiento

    Primera Tarea
    -> tarea_1

    Segunda Tarea
    -> tarea_2

    Tercera Tarea
    -> tarea_3

    Cuarta Tarea
    -> tarea_4

    Quinta Tarea
    -> tarea_5

    Sexta Tarea
    -> tarea_6

    Septima Tarea
    -> tarea_7


    ============================================================
    PRESENCIA DE LAS TAREAS
    ============================================================

    Primero determina si la tarea realmente existe.

    Si una tarea optativa está vacía:

    devuelve "".


    ============================================================
    PROPÓSITO PRINCIPAL DE LAS TAREAS
    ============================================================

    Las tareas deben analizarse especialmente desde esta pregunta:

    "¿Qué conocimientos base y competencias esenciales se deben
    reforzar para que el estudiante pueda afrontar con mayor
    solidez los contenidos del curso?"

    Por ello NO basta con comprobar que una tarea tenga relación
    general con el curso.

    Debes determinar qué aporta.


    ============================================================
    PARA CADA TAREA EXISTENTE
    ============================================================

    Analiza:

    - competencia;
    - título;
    - tecnologías;
    - descripción;
    - ponderación;
    - horas.

    Pero además determina:

    1. Qué conocimiento base pretende reforzar.

    2. Por qué ese conocimiento es necesario.

    3. Qué contenido posterior del curso se beneficia de ese
    refuerzo.

    4. Si la actividad realmente ejercita ese conocimiento.

    5. Si la competencia seleccionada corresponde.

    6. Si la descripción tiene sentido respecto del objetivo
    de refuerzo.

    7. Si herramientas y tecnologías son pertinentes.

    8. Si las horas son razonables.

    9. Si la ponderación tiene sentido.


    ============================================================
    TAREA DE FORTALECIMIENTO ACADÉMICO
    ============================================================

    Para:

    tarea_fortalecimiento

    debes prestar especial atención a:

    "Qué conocimientos base y competencias esenciales se deben
    reforzar que permitan al estudiante afrontar con mayor
    solidez los contenidos del curso."

    Comprueba:

    - si son conocimientos realmente previos;
    - si son necesarios;
    - si corresponden con dificultades previsibles;
    - si ayudan a los contenidos posteriores;
    - si están relacionados con el curso;
    - si la tarea propuesta realmente los refuerza.

    No aceptes una lista genérica de conocimientos solamente
    porque pertenezcan al área informática.


    ============================================================
    TAREAS NUMERADAS
    ============================================================

    Para cada tarea numerada comprueba también la progresión.

    Pregúntate:

    "¿Tiene sentido pedir esta tarea en este punto del curso?"

    Busca:

    - conocimientos todavía no enseñados;
    - tareas demasiado tardías;
    - tareas sin conexión con contenidos posteriores;
    - tareas que repiten actividades sin aportar refuerzo;
    - tareas cuyo título y descripción no coinciden;
    - tareas cuya competencia no corresponde.


    ============================================================
    NO CONFUNDIR AUSENCIA CON ERROR
    ============================================================

    Proyecto, práctica o tarea OPTATIVA sin contenido:

    ""

    No escribas:

    "Está vacío."
    "No fue definido."
    "Se encuentra configurado como optativo."
    "No tiene requerimientos."

    Simplemente devuelve "".

    Para componentes OBLIGATORIOS ausentes, sí corresponde generar
    una observación.


    ============================================================
    RESUMEN GENERAL
    ============================================================

    El resumen general debe sintetizar únicamente los principales
    hallazgos curriculares.

    Debe hablar de:

    - coherencia global;
    - problemas transversales;
    - competencias;
    - progresión;
    - tiempos cuando sean relevantes;
    - relación entre teoría y práctica;
    - relación con el contexto.

    NO incluyas:

    - filas;
    - celdas;
    - coordenadas;
    - detalles técnicos del backend;
    - cantidad de caracteres;
    - cantidad de hojas;
    - método de extracción;
    - número de tokens.

    NO uses:

    "fortalezas"
    "debilidades"

    como categorías.


    ============================================================
    ADVERTENCIAS
    ============================================================

    El arreglo "advertencias" debe utilizarse únicamente para
    situaciones que afecten la CONFIABILIDAD DEL ANÁLISIS.

    Ejemplos:

    - falta una fuente contextual necesaria;
    - una sección obligatoria no pudo ser interpretada;
    - existe información contradictoria que impide determinar
    una conclusión.

    NO coloques aquí todos los errores curriculares.

    Los errores curriculares pertenecen a "observaciones".


    ============================================================
    SECCIONES DINÁMICAS DE LA MATRIZ
    ============================================================

    Debes devolver EXACTAMENTE estas claves:

    {target_text}

    No inventes otras claves.

    No elimines una clave requerida por el esquema.

    Para un proyecto, práctica o tarea optativa sin contenido:

    usa "".


    ============================================================
    DOCUMENTOS DISPONIBLES
    ============================================================

    A continuación se proporciona el contenido extraído de los
    documentos.

    IMPORTANTE:

    Los identificadores técnicos que puedan acompañar al texto
    (coordenadas, posiciones u otros datos de extracción) sirven
    solo para reconstruir la estructura.

    NO deben aparecer en tu respuesta.

    {context_text}


    ============================================================
    FORMATO JSON OBLIGATORIO
    ============================================================

    Responde únicamente con:

    {{
    "observaciones": {{
    {example_fields}
    }},
    "resumen_general": "síntesis crítica de la coherencia global del diseño curricular",
    "advertencias": []
    }}
    """.strip()

        return (
            system_message,
            user_message,
        )

    # ========================================================
    # DEEPSEEK
    # ========================================================

    def _call_deepseek(
        self,
        system_message: str,
        user_message: str,
    ) -> Dict[str, Any]:

        api_key = (
            getattr(
                settings,
                "DEEPSEEK_API_KEY",
                None,
            )
            or ""
        ).strip()

        if not api_key:

            raise CurriculumFeedbackError(
                "DEEPSEEK_API_KEY no está configurada"
            )

        payload = json.dumps(
            {
                "model": DEEPSEEK_MODEL,

                "messages": [
                    {
                        "role": "system",
                        "content": system_message,
                    },
                    {
                        "role": "user",
                        "content": user_message,
                    },
                ],

                # Para esta tarea queremos salida estructurada
                # y no necesitamos exponer razonamiento interno.
                "thinking": {
                    "type": "disabled",
                },

                "temperature": 0.1,

                "max_tokens": (
                    DEEPSEEK_MAX_OUTPUT_TOKENS
                ),

                "response_format": {
                    "type": "json_object",
                },
            },
            ensure_ascii=False,
        ).encode(
            "utf-8"
        )

        last_error: Optional[
            Exception
        ] = None

        for attempt in range(
            1,
            DEEPSEEK_MAX_ATTEMPTS + 1,
        ):

            print(
                "🤖 [CURRICULUM] DeepSeek | "
                f"intento={attempt}/"
                f"{DEEPSEEK_MAX_ATTEMPTS} | "
                f"modelo={DEEPSEEK_MODEL} | "
                f"request_bytes={len(payload)}",
                flush=True,
            )

            request = (
                urllib.request.Request(
                    DEEPSEEK_URL,
                    data=payload,
                    headers={
                        "Authorization": (
                            f"Bearer {api_key}"
                        ),
                        "Content-Type": (
                            "application/json"
                        ),
                    },
                    method="POST",
                )
            )

            try:

                with urllib.request.urlopen(
                    request,
                    timeout=(
                        DEEPSEEK_TIMEOUT_SECONDS
                    ),
                ) as response:

                    response_body = (
                        response
                        .read()
                        .decode(
                            "utf-8"
                        )
                    )

                response_json = (
                    json.loads(
                        response_body
                    )
                )

                choices = (
                    response_json
                    .get(
                        "choices",
                        [],
                    )
                )

                if not choices:

                    raise (
                        CurriculumFeedbackError(
                            "DeepSeek no devolvió "
                            "ninguna respuesta"
                        )
                    )

                choice = choices[0]

                finish_reason = (
                    choice.get(
                        "finish_reason"
                    )
                )

                if (
                    finish_reason
                    == "length"
                ):

                    raise (
                        CurriculumFeedbackError(
                            "DeepSeek cortó la "
                            "respuesta por límite "
                            "de salida"
                        )
                    )

                raw = (
                    choice
                    .get(
                        "message",
                        {},
                    )
                    .get(
                        "content",
                        "",
                    )
                    .strip()
                )

                # Defensa adicional si el modelo envía
                # ```json ... ```
                raw = re.sub(
                    r"^```json?\s*",
                    "",
                    raw,
                    flags=re.IGNORECASE,
                )

                raw = re.sub(
                    r"\s*```$",
                    "",
                    raw,
                )

                parsed = json.loads(
                    raw
                )

                return {
                    "data": parsed,

                    "usage": (
                        response_json
                        .get(
                            "usage",
                            {},
                        )
                    ),

                    "model": (
                        response_json
                        .get(
                            "model",
                            DEEPSEEK_MODEL,
                        )
                    ),
                }

            except urllib.error.HTTPError as exc:

                body = (
                    exc.read()
                    .decode(
                        "utf-8",
                        errors="replace",
                    )
                )

                last_error = (
                    CurriculumFeedbackError(
                        f"DeepSeek HTTP "
                        f"{exc.code}: "
                        f"{body[:1000]}"
                    )
                )

                # Solo reintentar errores transitorios.
                if exc.code not in {
                    408,
                    409,
                    429,
                    500,
                    502,
                    503,
                    504,
                }:
                    break

            except (
                urllib.error.URLError,
                TimeoutError,
            ) as exc:

                last_error = exc

            except json.JSONDecodeError as exc:

                last_error = (
                    CurriculumFeedbackError(
                        "DeepSeek devolvió "
                        "JSON inválido: "
                        f"{exc}"
                    )
                )

                break

            except CurriculumFeedbackError:
                raise

            except Exception as exc:

                last_error = exc

            if (
                attempt
                < DEEPSEEK_MAX_ATTEMPTS
            ):

                wait_seconds = min(
                    4,
                    2 ** (
                        attempt - 1
                    ),
                )

                time.sleep(
                    wait_seconds
                )

        raise CurriculumFeedbackError(
            "No se pudo completar el análisis "
            f"con DeepSeek: {last_error}"
        )

    # ========================================================
    # VALIDAR JSON DE IA
    # ========================================================

    def _validate_feedback_payload(
        self,
        payload: Dict[str, Any],
        matrix_targets: List[
            Dict[str, Any]
        ],
    ) -> Dict[str, Any]:

        observations = payload.get(
            "observaciones"
        )

        if not isinstance(
            observations,
            dict,
        ):

            raise CurriculumFeedbackError(
                "La IA no devolvió el objeto "
                "'observaciones'"
            )

        expected_ids = [
            target["id"]
            for target
            in matrix_targets
        ]

        missing = [
            target_id
            for target_id
            in expected_ids
            if target_id
            not in observations
        ]

        if missing:

            raise CurriculumFeedbackError(
                "La respuesta de IA está incompleta. "
                "Faltan: "
                + ", ".join(
                    missing
                )
            )

        clean = {
            target_id: str(
                observations.get(
                    target_id
                )
                or ""
            ).strip()

            for target_id
            in expected_ids
        }

        ai_warnings = payload.get(
            "advertencias"
        )

        if not isinstance(
            ai_warnings,
            list,
        ):
            ai_warnings = []

        return {
            "observaciones": clean,

            "resumen_general": str(
                payload.get(
                    "resumen_general"
                )
                or ""
            ).strip(),

            "advertencias": [
                str(item).strip()
                for item
                in ai_warnings
                if str(item).strip()
            ],
        }
        
        
    # ========================================================
    # ESCRIBIR ARCHIVO 6
    # ========================================================

    def write_matrix_feedback(
        self,
        spreadsheet_id: str,
        sheet_title: str,
        matrix_targets: List[
            Dict[str, Any]
        ],
        observations: Dict[
            str,
            str
        ],
    ) -> Dict[str, Any]:

        escaped_title = (
            sheet_title.replace(
                "'",
                "''",
            )
        )

        updates: List[
            Dict[str, Any]
        ] = []

        for target in matrix_targets:

            target_id = target[
                "id"
            ]

            row_number = target[
                "row"
            ]

            # G es Observaciones Generales por la IA.
            cell_range = (
                f"'{escaped_title}'!"
                f"{MATRIX_AI_COLUMN}"
                f"{row_number}"
            )

            updates.append(
                {
                    "range": cell_range,

                    # También escribimos "".
                    # Así se eliminan observaciones antiguas
                    # de prácticas/tareas que ya no existan.
                    "values": [
                        [
                            observations.get(
                                target_id,
                                "",
                            )
                        ]
                    ],
                }
            )

        result = (
            google_sheets_service
            .batch_update_values(
                spreadsheet_id,
                updates,
            )
        )

        updated_cells = (
            result.get(
                "updated"
            )
            if isinstance(
                result,
                dict,
            )
            else None
        )

        if updated_cells is None:
            updated_cells = result.get(
                "totalUpdatedCells",
                0,
            )

        return {
            "success": True,
            "sheet": sheet_title,
            "column": MATRIX_AI_COLUMN,
            "updated": updated_cells,
            "targets": len(
                matrix_targets
            ),
        }

    # ========================================================
    # ANALIZAR UN CURSO
    # ========================================================

    def analyze_course(
        self,
        course: CourseCatalog,
        semester: str,
        year: int,
        write_output: bool = False,
    ) -> Dict[str, Any]:
        """
        Análisis completo de UN curso.

        Este método será reutilizado posteriormente por el
        background worker para procesar todos los cursos.
        """
        started_at = (
            time.perf_counter()
        )

        print(
            "\n"
            "=========================================\n"
            "🎓 RETROALIMENTACIÓN CURRICULAR\n"
            f"Curso: {course.code} - {course.name}\n"
            f"Semestre: {semester}\n"
            f"Año: {year}\n"
            "=========================================",
            flush=True,
        )

        # ----------------------------------------------------
        # 1. Localizar
        # ----------------------------------------------------

        locations = self.locate_course(
            course,
            semester,
            year,
        )
        
        revision_folder = (
            self._find_revision_folder(
                locations[
                    "course"
                ]["id"]
            )
        )

        if not revision_folder:

            raise CurriculumFeedbackError(
                "No se encontró "
                "0_Revision_de_Material"
            )

        matrix_file = (
            self._find_matrix_file(
                revision_folder[
                    "id"
                ]
            )
        )

        if not matrix_file:

            raise CurriculumFeedbackError(
                "No se encontró "
                "02_Matriz observaciones estructura"
            )

        # ----------------------------------------------------
        # 2. Archivos
        # ----------------------------------------------------

        files = self.locate_planning_files(
            locations[
                "planning"
            ]["id"]
        )
        
        if not files.get(
            MAIN_INPUT_KEY
        ):

            raise CurriculumFeedbackError(
                "No se encontró "
                "5_Diseño_Curricular"
            )

        # ----------------------------------------------------
        # 3. Cargar PDF/XLSX
        # ----------------------------------------------------

        documents, warnings = (
            self.load_context(
                files
            )
        )

        # ----------------------------------------------------
        # 4. Construir contexto
        # ----------------------------------------------------

        (
            context_text,
            context_truncated,
        ) = self._context_text(
            documents
        )
        
        (
            matrix_sheet_title,
            matrix_targets,
        ) = self._matrix_targets(
            matrix_file[
                "id"
            ]
        )

        if context_truncated:

            warnings.append(
                "El contexto superó el límite interno "
                "de seguridad del backend y fue recortado. "
                "Se priorizó 5_Diseño_Curricular."
            )

        # ----------------------------------------------------
        # 5. Prompt
        # ----------------------------------------------------

        (
            system_message,
            user_message,
        ) = self._build_prompt(
            course,
            context_text,
            warnings,
            matrix_targets,
        )

        response_schema = (
            self._matrix_response_schema(
                matrix_targets
            )
        )

        # ----------------------------------------------------
        # 6. DeepSeek
        # ----------------------------------------------------

        ai_result = self._call_ai(
            system_message,
            user_message,
            response_schema=(
                response_schema
            ),
        )

        print(
            "✅ [CURRICULUM] IA terminada correctamente | "
            f"provider={ai_result.get('provider')} | "
            f"model={ai_result.get('model')}",
            flush=True,
        )

        print(
            "🔎 [CURRICULUM] Validando respuesta estructurada...",
            flush=True,
        )

        # ----------------------------------------------------
        # 7. Validar
        # ----------------------------------------------------

        validated = (
            self._validate_feedback_payload(
                ai_result[
                    "data"
                ],
                matrix_targets,
            )
        )
        
        print(
            "✅ [CURRICULUM] JSON de observaciones válido | "
            f"apartados={len(validated.get('observaciones', {}))}",
            flush=True,
        )

        # ----------------------------------------------------
        # 8. Escribir opcionalmente
        # ----------------------------------------------------

        write_result = None

        if write_output:

            write_result = (
                self.write_matrix_feedback(
                    spreadsheet_id=(
                        matrix_file[
                            "id"
                        ]
                    ),

                    sheet_title=(
                        matrix_sheet_title
                    ),

                    matrix_targets=(
                        matrix_targets
                    ),

                    observations=(
                        validated[
                            "observaciones"
                        ]
                    ),
                )
            )

        # ----------------------------------------------------
        # 9. Resultado
        # ----------------------------------------------------

        result = {
            "success": True,

            "course": {
                "code": str(
                    course.code
                ),

                "name": (
                    course.name
                ),

                "area": (
                    course_contacts_service
                    .canonical_area(
                        course.area
                    )
                ),

                "semester": (
                    course_contacts_service
                    .normalize_semester(
                        semester
                    )
                ),

                "year": year,
            },

            "warnings": warnings,

            "documents": {
                key: {
                    "name": value[
                        "name"
                    ],

                    "chars": value[
                        "chars"
                    ],

                    "extraction_method": (
                        value[
                            "extraction_method"
                        ]
                    ),
                }

                for key, value
                in documents.items()
            },

            "context_chars_sent": len(
                context_text
            ),

            "context_truncated": (
                context_truncated
            ),

            "provider": {
                "name": ai_result.get(
                    "provider",
                    "unknown",
                ),

                "model": ai_result.get(
                    "model"
                ),

                "usage": ai_result.get(
                    "usage",
                    {},
                ),
            },

            "result": validated,

            "write": write_result,

            "elapsed_seconds": round(
                time.perf_counter()
                - started_at,
                2,
            ),
            "matrix": {
                "id": matrix_file.get(
                    "id"
                ),

                "name": matrix_file.get(
                    "name"
                ),

                "webViewLink": (
                    matrix_file.get(
                        "webViewLink"
                    )
                ),

                "sheet": (
                    matrix_sheet_title
                ),

                "column": (
                    MATRIX_AI_COLUMN
                ),
            },

            "matrix_targets": (
                matrix_targets
            ),
        }

        print(
            "🏁 [CURRICULUM] Curso completamente procesado | "
            f"curso={course.code} | "
            f"segundos={result['elapsed_seconds']} | "
            f"write_output={write_output}",
            flush=True,
        )

        return result


curriculum_feedback_service = (
    CurriculumFeedbackService()
)